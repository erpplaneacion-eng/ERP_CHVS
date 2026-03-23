"""
Utilidades de dibujo para reportes de análisis nutricional en Excel.

Este módulo contiene una clase base con lógica reutilizable para dibujar
las diferentes secciones de un reporte de análisis nutricional.
"""

import io
import os
import urllib.request
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
from urllib.parse import urlparse

from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# =====================================================================
# CONSTANTES Y CLASES DE DATOS REUTILIZABLES
# =====================================================================

@dataclass(frozen=True)
class ExcelLayout:
    """Constantes de diseño del Excel (PAE - 5 niveles educativos)"""
    # Filas
    TITLE_ROW = 1
    ADMIN_START_ROW = 2
    ADMIN_END_ROW = 8
    HEADER_START_ROW = 10
    HEADER_END_ROW = 10
    DATA_START_ROW = 11

    # Columnas
    COL_COMPONENTE = 1
    COL_GRUPO = 2
    COL_PREPARACION = 3
    COL_INGREDIENTE = 4
    COL_CODIGO = 5
    COL_PESO_BRUTO = 6
    COL_PESO_NETO = 7
    COL_CALORIAS = 8
    COL_PROTEINA = 9
    COL_GRASA = 10
    COL_CHO = 11
    COL_CALCIO = 12
    COL_HIERRO = 13
    COL_SODIO = 14

    # Rangos
    TITLE_RANGE = 'D1:N1'
    TOTAL_LABEL_COLS = 3  # A:C
    LABEL_MERGE_END = 7   # Merge A:G para etiquetas de totales/adecuación
    LAST_DATA_COL = 14    # Última columna con datos (SODIO)
    TITLE_END_COL = 14    # Fin del merge del título


@dataclass(frozen=True)
class ComedoresExcelLayout:
    """Constantes de diseño para Comedores / Adulto Mayor (sin niveles, 4 nutrientes).
    No hay columna GRUPO; las columnas se desplazan una posición a la izquierda y
    los nutrientes empiezan en la columna 7.
    """
    # Columnas
    COL_COMPONENTE: int = 1
    COL_PREPARACION: int = 2   # Sin columna GRUPO
    COL_INGREDIENTE: int = 3
    COL_CODIGO: int = 4
    COL_PESO_BRUTO: int = 5
    COL_PESO_NETO: int = 6
    COL_CALORIAS: int = 7      # Primera columna de nutrientes
    COL_PROTEINA: int = 8
    COL_GRASA: int = 9
    COL_CHO: int = 10

    # Rangos
    LABEL_MERGE_END: int = 6   # Merge A:F para etiquetas de totales/adecuación
    LAST_DATA_COL: int = 10    # Última columna con datos (CHO)
    TITLE_END_COL: int = 10    # Fin del merge del título


@dataclass(frozen=True)
class ExcelStyles:
    """Estilos predefinidos para el Excel"""
    HEADER_COLOR = "B8D4F0"
    TOTAL_COLOR = "E6F3FF"

    @staticmethod
    def get_header_fill() -> PatternFill:
        return PatternFill(
            start_color=ExcelStyles.HEADER_COLOR,
            end_color=ExcelStyles.HEADER_COLOR,
            fill_type="solid"
        )

    @staticmethod
    def get_total_fill() -> PatternFill:
        return PatternFill(
            start_color=ExcelStyles.TOTAL_COLOR,
            end_color=ExcelStyles.TOTAL_COLOR,
            fill_type="solid"
        )

    @staticmethod
    def get_border() -> Border:
        side = Side(style='thin')
        return Border(left=side, right=side, top=side, bottom=side)


@dataclass(frozen=True)
class ColumnWidths:
    """Anchos de columnas"""
    WIDTHS = {
        'A': 25, 'B': 25, 'C': 15, 'D': 35, 'E': 15, 'F': 25,
        'G': 15, 'H': 12, 'I': 12, 'J': 12, 'K': 10, 'L': 10,
        'M': 10, 'N': 10, 'O': 12, 'P': 12
    }


NUTRITIONAL_HEADERS = [
    "COMPONENTE",
    "GRUPO DE ALIMENTOS",
    "NOMBRE DE LA PREPARACION",
    "NOMBRE DEL ALIMENTO (Ingredientes)",
    "CÓDIGO DEL ALIMENTO",
    "PESO BRUTO (g)",
    "PESO NETO (g)",
    "CALORÍAS (Kcal)",
    "PROTEÍNA (g)",
    "GRASA (g)",
    "CHO (g)",
    "CALCIO (mg)",
    "HIERRO (mg)",
    "SODIO (mg)"
]

# Encabezados para Comedores Comunitarios / Adulto Mayor (sin GRUPO, 4 nutrientes, 10 cols)
COMEDORES_NUTRITIONAL_HEADERS = [
    "COMPONENTES",
    "NOMBRE DE LA PREPARACION",
    "NOMBRE DEL ALIMENTO (Ingredientes)",
    "CÓDIGO DEL ALIMENTO",
    "PESO BRUTO (g - cc)",
    "PESO NETO (g - cc)",
    "ENERGÍA (Kcal)",
    "PROTEÍNA (g)",
    "GRASA (g)",
    "CHO (g)",
]

NUTRIENT_KEYS = [
    'calorias', 'proteina', 'grasa', 'cho',
    'calcio', 'hierro', 'sodio'
]

# Sólo 4 nutrientes para Comedores (no evalúan calcio, hierro, sodio)
NUTRIENT_KEYS_COMEDORES = ['calorias', 'proteina', 'grasa', 'cho']


# =====================================================================
# CLASE BASE PARA DIBUJAR REPORTES
# =====================================================================

class ExcelReportDrawer:
    """
    Clase base que contiene métodos reutilizables para dibujar un análisis
    nutricional en una hoja de cálculo de openpyxl.
    """

    def __init__(self):
        self.layout = ExcelLayout()
        self.comedores_layout = ComedoresExcelLayout()
        self.styles = ExcelStyles()
        self.header_fill = self.styles.get_header_fill()
        self.total_fill = self.styles.get_total_fill()
        self.border = self.styles.get_border()

    def _insert_logo(self, ws: Worksheet, logo_path: str, start_cell: str = 'A1') -> None:
        """Inserta el logo del programa en la celda especificada."""
        if not logo_path:
            return

        try:
            source = str(logo_path).strip()
            parsed = urlparse(source)

            # URL remota (Cloudinary en producción): descargar a BytesIO.
            # openpyxl 3.1+ almacena self._buf en memoria desde __init__,
            # por lo que no necesita que el archivo exista al guardar el workbook.
            if parsed.scheme in ('http', 'https'):
                with urllib.request.urlopen(source) as resp:
                    img_source = io.BytesIO(resp.read())
            # Ruta web local (/media/...): resolver a MEDIA_ROOT si es posible.
            elif source.startswith('/'):
                img_source = source
                try:
                    from django.conf import settings
                    media_root = getattr(settings, 'MEDIA_ROOT', '') or ''
                    media_url = getattr(settings, 'MEDIA_URL', '/media/') or '/media/'
                    if source.startswith(media_url) and media_root:
                        relative = source[len(media_url):].lstrip('/\\')
                        candidate = os.path.join(media_root, relative)
                        if os.path.exists(candidate):
                            img_source = candidate
                except Exception:
                    pass
            else:
                img_source = source

            img = Image(img_source)
            img.height = 60
            img.width = 200
            ws.add_image(img, start_cell)
            ws.row_dimensions[int(start_cell[1:])].height = 50
        except Exception as e:
            print(f"Warning: no se pudo insertar logo '{logo_path}': {e}")

    def _populate_title(self, ws: Worksheet, start_row: int) -> None:
        """Poblar título principal en una fila específica."""
        title_cell = ws.cell(row=start_row, column=4)
        title_cell.value = "ANÁLISIS NUTRICIONAL"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

    def _populate_administrative_section(
        self,
        ws: Worksheet,
        start_row: int,
        analisis_data: Dict,
        nivel_escolar_nombre: Optional[str] = None
    ) -> None:
        """Poblar información administrativa del menú."""
        menu_info = analisis_data.get('menu', {})
        nivel_escolar = nivel_escolar_nombre or self._extract_nivel_escolar(analisis_data)
        modalidad_id = menu_info.get('modalidad_id')
        modalidad_nombre = menu_info.get('modalidad', 'N/A')
        modalidad_atencion = self._resolver_modalidad_atencion(modalidad_id)

        def _upper(value):
            return str(value).upper() if value is not None else ""

        admin_data = [
            ("ENTIDAD TERRITORIAL:", _upper(menu_info.get('programa', 'N/A'))),
            ("MINUTA CON ENFOQUE ETNICO:", _upper("No")),
            ("GRUPO ÉTNICO", _upper("Sin Pertenencia Étnica")),
            ("MODALIDAD DE ATENCIÓN", _upper(modalidad_atencion)),
            ("TIPO DE COMPLEMENTO", _upper(modalidad_nombre)),
            ("NIVEL", _upper(nivel_escolar)),
            ("MENÚ No.", _upper(menu_info.get('nombre', 'N/A')))
        ]

        for i, (label, value) in enumerate(admin_data):
            row = start_row + i
            label_cell = ws.cell(row=row, column=1)
            label_cell.value = label
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(horizontal='left', vertical='center')

            value_cell = ws.cell(row=row, column=4)

            # Debug: verificar tipo de celda
            cell_type = type(value_cell).__name__
            if cell_type != 'Cell':
                print(f"[WARNING] Celda en fila {row}, columna 4 es tipo: {cell_type}")
                print(f"[WARNING] Merged cells en hoja: {ws.merged_cells}")
                # Intentar unmerge si es necesario
                from openpyxl.worksheet.cell_range import CellRange
                for merged_range in list(ws.merged_cells.ranges):
                    if value_cell.coordinate in merged_range:
                        print(f"[WARNING] Deshaciendo merge: {merged_range}")
                        ws.unmerge_cells(str(merged_range))
                        value_cell = ws.cell(row=row, column=4)  # Obtener celda de nuevo
                        break

            value_cell.value = value
            value_cell.alignment = Alignment(horizontal='left', vertical='center')

    @staticmethod
    def _resolver_modalidad_atencion(modalidad_id: Optional[str]) -> str:
        modalidad_codigo = str(modalidad_id or '').strip()
        if modalidad_codigo in {'020511', '20502'}:
            return "Ración Industrializada"
        return "Ración para Preparar en Sitio"

    def _populate_headers(self, ws: Worksheet, start_row: int, is_comedores: bool = False) -> None:
        """Poblar encabezados de columnas."""
        headers = COMEDORES_NUTRITIONAL_HEADERS if is_comedores else NUTRITIONAL_HEADERS
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border

    def _populate_ingredients_from_analysis(self, ws: Worksheet, start_row: int, nivel_data: Dict, is_comedores: bool = False) -> int:
        """Poblar ingredientes desde datos de análisis."""
        current_row = start_row
        left_align = Alignment(vertical='center', horizontal='left', wrap_text=True)

        # Layout dinámico según tipo de programa
        layout = self.comedores_layout if is_comedores else self.layout

        # Contador de ingredientes totales procesados
        total_ingredients_added = 0

        for preparacion in nivel_data.get('preparaciones', []):
            start_row_for_prep = current_row
            ingredients_in_prep = [ing for ing in preparacion.get('ingredientes', []) if ing.get('alimento_encontrado')]
            num_ingredients = len(ingredients_in_prep)

            if num_ingredients == 0:
                continue

            total_ingredients_added += num_ingredients

            ws.cell(row=start_row_for_prep, column=layout.COL_PREPARACION).value = preparacion.get('nombre', '')
            ws.cell(row=start_row_for_prep, column=layout.COL_PREPARACION).alignment = left_align

            # Para agrupar componentes consecutivos
            componente_groups = []  # List of [componente_str, start_row, end_row]

            for ingrediente in ingredients_in_prep:
                comp_ing = str(ingrediente.get('componente', '')).strip()
                if not comp_ing:
                    comp_ing = str(preparacion.get('componente', 'SIN COMPONENTE')).strip()

                # Agrupar componentes consecutivos iguales
                if componente_groups and componente_groups[-1][0] == comp_ing:
                    componente_groups[-1][2] = current_row
                else:
                    componente_groups.append([comp_ing, current_row, current_row])

                # Columna GRUPO: solo en PAE
                if not is_comedores:
                    ws.cell(row=current_row, column=self.layout.COL_GRUPO).value = ingrediente.get(
                        'grupo_alimentos',
                        preparacion.get('grupo_alimentos', 'SIN GRUPO')
                    )

                ws.cell(row=current_row, column=layout.COL_INGREDIENTE).value = ingrediente.get('nombre', '')
                ws.cell(row=current_row, column=layout.COL_CODIGO).value = ingrediente.get('codigo_icbf', '')
                ws.cell(row=current_row, column=layout.COL_PESO_BRUTO).value = ingrediente.get('peso_bruto_base', 0)
                ws.cell(row=current_row, column=layout.COL_PESO_NETO).value = ingrediente.get('peso_neto_base', 0)

                valores = ingrediente.get('valores_finales_guardados')
                if not valores:
                    peso_neto = float(ingrediente.get('peso_neto_base', 0) or 0)
                    valores_por_100g = ingrediente.get('valores_por_100g', {}) or {}
                    factor = peso_neto / 100
                    valores = {
                        'calorias': (valores_por_100g.get('calorias_kcal', 0) or 0) * factor,
                        'proteina': (valores_por_100g.get('proteina_g', 0) or 0) * factor,
                        'grasa': (valores_por_100g.get('grasa_g', 0) or 0) * factor,
                        'cho': (valores_por_100g.get('cho_g', 0) or 0) * factor,
                        'calcio': (valores_por_100g.get('calcio_mg', 0) or 0) * factor,
                        'hierro': (valores_por_100g.get('hierro_mg', 0) or 0) * factor,
                        'sodio': (valores_por_100g.get('sodio_mg', 0) or 0) * factor,
                    }

                ws.cell(row=current_row, column=layout.COL_CALORIAS).value = valores.get('calorias', 0)
                ws.cell(row=current_row, column=layout.COL_PROTEINA).value = valores.get('proteina', 0)
                ws.cell(row=current_row, column=layout.COL_GRASA).value = valores.get('grasa', 0)
                ws.cell(row=current_row, column=layout.COL_CHO).value = valores.get('cho', 0)
                if not is_comedores:
                    ws.cell(row=current_row, column=self.layout.COL_CALCIO).value = valores.get('calcio', 0)
                    ws.cell(row=current_row, column=self.layout.COL_HIERRO).value = valores.get('hierro', 0)
                    ws.cell(row=current_row, column=self.layout.COL_SODIO).value = valores.get('sodio', 0)

                # Bordes hasta la última columna de datos
                for col_idx in range(layout.COL_COMPONENTE, layout.LAST_DATA_COL + 1):
                    ws.cell(row=current_row, column=col_idx).border = self.border

                current_row += 1

            # Renderizar componentes y mergear si hay más de una fila
            for comp_str, start_r, end_r in componente_groups:
                cell = ws.cell(row=start_r, column=layout.COL_COMPONENTE)
                cell.value = comp_str.upper() if is_comedores else comp_str
                cell.alignment = left_align
                if end_r > start_r:
                    ws.merge_cells(
                        start_row=start_r, start_column=layout.COL_COMPONENTE,
                        end_row=end_r, end_column=layout.COL_COMPONENTE
                    )

            if num_ingredients > 1:
                end_row_for_prep = current_row - 1
                cell_to_align = ws.cell(row=start_row_for_prep, column=layout.COL_PREPARACION)
                cell_to_align.alignment = left_align
                ws.merge_cells(
                    start_row=start_row_for_prep, start_column=layout.COL_PREPARACION,
                    end_row=end_row_for_prep, end_column=layout.COL_PREPARACION
                )

        # Si no se agregó ningún ingrediente, retornar start_row - 1
        if total_ingredients_added == 0:
            return start_row - 1

        return current_row - 1

    def _add_calculations_section(self, ws: Worksheet, last_data_row: int, nivel_data: Dict, is_comedores: bool = False) -> None:
        """Agregar sección de totales, recomendaciones y adecuación."""
        total_row = last_data_row + 2
        self._add_totals_row(ws, total_row, nivel_data, is_comedores)
        req_row = total_row + 1
        self._add_recommendations_row(ws, req_row, nivel_data, is_comedores)
        pct_row = req_row + 1
        self._add_adequacy_row(ws, pct_row, total_row, req_row, nivel_data, is_comedores)

    def _add_totals_row(self, ws: Worksheet, row: int, nivel_data: Dict, is_comedores: bool = False) -> None:
        """Agregar fila de totales."""
        layout = self.comedores_layout if is_comedores else self.layout
        label = "APORTE TOTAL NUTRIENTES DEL MENÚ" if is_comedores else "TOTAL MENÚ"

        ws.cell(row, 1).value = label
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 1).fill = self.total_fill
        ws.cell(row, 1).border = self.border
        ws.cell(row, 1).alignment = Alignment(horizontal='center')
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=layout.LABEL_MERGE_END
        )

        keys = NUTRIENT_KEYS_COMEDORES if is_comedores else NUTRIENT_KEYS
        totales = nivel_data.get('totales', {})
        for i, key in enumerate(keys, start=layout.COL_CALORIAS):
            cell = ws.cell(row, i)
            cell.value = totales.get(key, 0)
            cell.fill = self.total_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')

    def _add_recommendations_row(self, ws: Worksheet, row: int, nivel_data: Dict, is_comedores: bool = False) -> None:
        """Agregar fila de recomendaciones."""
        layout = self.comedores_layout if is_comedores else self.layout

        label_cell = ws.cell(row, 1)
        if is_comedores:
            label_cell.value = "RECOMENDACIÓN PARA EL GRUPO DE EDAD"
            label_cell.font = Font(bold=True)
        label_cell.border = self.border
        label_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=layout.LABEL_MERGE_END
        )

        requerimientos = nivel_data.get('requerimientos', {})
        default_values = {
            'calorias': 1300,
            'proteina': 45.5,
            'grasa': 43.3,
            'cho': 182,
            'calcio': 700,
            'hierro': 5.6,
            'sodio': 1133,
        }
        keys = NUTRIENT_KEYS_COMEDORES if is_comedores else NUTRIENT_KEYS
        for i, key in enumerate(keys, start=layout.COL_CALORIAS):
            cell = ws.cell(row, i)
            val = requerimientos.get(key)
            cell.value = val if val is not None else default_values.get(key, 0)
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')

    def _add_adequacy_row(self, ws: Worksheet, row: int, total_row: int, req_row: int, nivel_data: Dict = None, is_comedores: bool = False) -> None:
        """Agregar fila de % adecuación.
        Calcula los valores directamente en Python para garantizar que se muestren en Excel.
        """
        layout = self.comedores_layout if is_comedores else self.layout
        label = "% CUBRIMIENTO DEL MENÚ" if is_comedores else "% ADECUACIÓN"

        ws.cell(row, 1).value = label
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 1).border = self.border
        ws.cell(row, 1).alignment = Alignment(horizontal='center')
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=layout.LABEL_MERGE_END
        )

        keys = NUTRIENT_KEYS_COMEDORES if is_comedores else NUTRIENT_KEYS
        totales = (nivel_data or {}).get('totales', {}) or {}
        requerimientos = (nivel_data or {}).get('requerimientos', {}) or {}

        for i, key in enumerate(keys, start=layout.COL_CALORIAS):
            cell = ws.cell(row, i)
            total = float(totales.get(key) or 0)
            req = float(requerimientos.get(key) or 0)
            value = round(total / req * 100, 2) if req > 0 else 0
            cell.value = value
            cell.border = self.border
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal='center')

    def _add_signatures(self, ws: Worksheet, start_row: int, menu_info: Optional[Dict] = None) -> None:
        """Agregar sección de firmas."""
        bottom_border = Border(bottom=Side(style='thin'))
        full_border = self.border
        center_align = Alignment(horizontal='center', vertical='center')
        firmas = (menu_info or {}).get('firmas') or {}

        elabora_nombre = firmas.get('elabora_nombre') or "SARA ISABEL DIAZ MARQUEZ"
        elabora_matricula = firmas.get('elabora_matricula') or "1107089938"
        elabora_firma_texto = firmas.get('elabora_firma_texto') or ""
        elabora_firma_imagen_path = firmas.get('elabora_firma_imagen_path')

        aprueba_nombre = firmas.get('aprueba_nombre') or "GABRIELA GIRALDO MARTINEZ"
        aprueba_matricula = firmas.get('aprueba_matricula') or "1005964870"
        aprueba_firma_texto = firmas.get('aprueba_firma_texto') or ""
        aprueba_firma_imagen_path = firmas.get('aprueba_firma_imagen_path')

        row = start_row
        # --- Bloque 1 ---
        ws.cell(row=row, column=1).value = "NOMBRE NUTRICIONISTA - DIETISTA QUE ELABORA EL ANÁLISIS"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).border = full_border
        ws.merge_cells(f'A{row}:D{row}')

        name_value_cell = ws.cell(row=row, column=5)
        name_value_cell.value = elabora_nombre
        name_value_cell.border = full_border
        name_value_cell.alignment = center_align
        ws.merge_cells(f'E{row}:G{row}')

        row += 3
        ws.cell(row=row, column=1).value = "FIRMA"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).border = full_border
        
        for col_idx in range(2, 5):
            ws.cell(row=row, column=col_idx).border = bottom_border
        ws.merge_cells(f'B{row}:D{row}')
        if elabora_firma_imagen_path:
            self._insert_signature_image(ws, elabora_firma_imagen_path, f'B{max(1, row - 2)}')
        elif elabora_firma_texto:
            firma_cell = ws.cell(row=row, column=2)
            firma_cell.value = elabora_firma_texto
            firma_cell.alignment = center_align

        ws.cell(row=row, column=8).value = "MATRÍCULA PROFESIONAL"
        ws.cell(row=row, column=8).border = full_border
        ws.merge_cells(f'H{row}:J{row}')
        
        matricula_value_cell = ws.cell(row=row, column=11)
        matricula_value_cell.value = elabora_matricula
        matricula_value_cell.border = full_border
        matricula_value_cell.alignment = center_align
        ws.merge_cells(f'K{row}:L{row}')

        row += 2
        # --- Bloque 2 ---
        ws.cell(row=row, column=1).value = "NOMBRE NUTRICIONISTA - DIETISTA QUE REVISA Y APRUEBA EL ANÁLISIS POR PARTE DE LA SEM"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).border = full_border
        ws.merge_cells(f'A{row}:D{row}')

        name_value_cell_2 = ws.cell(row=row, column=5)
        name_value_cell_2.value = aprueba_nombre
        name_value_cell_2.border = full_border
        name_value_cell_2.alignment = center_align
        ws.merge_cells(f'E{row}:G{row}')

        row += 3
        ws.cell(row=row, column=1).value = "FIRMA"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).border = full_border

        for col_idx in range(2, 5):
            ws.cell(row=row, column=col_idx).border = bottom_border
        ws.merge_cells(f'B{row}:D{row}')
        if aprueba_firma_imagen_path:
            self._insert_signature_image(ws, aprueba_firma_imagen_path, f'B{max(1, row - 2)}')
        elif aprueba_firma_texto:
            firma_cell_2 = ws.cell(row=row, column=2)
            firma_cell_2.value = aprueba_firma_texto
            firma_cell_2.alignment = center_align

        ws.cell(row=row, column=8).value = "MATRÍCULA PROFESIONAL"
        ws.cell(row=row, column=8).border = full_border
        ws.merge_cells(f'H{row}:J{row}')

        matricula_value_cell_2 = ws.cell(row=row, column=11)
        matricula_value_cell_2.value = aprueba_matricula
        matricula_value_cell_2.border = full_border
        matricula_value_cell_2.alignment = center_align
        ws.merge_cells(f'K{row}:L{row}')

    def _insert_signature_image(self, ws: Worksheet, image_path: str, anchor_cell: str) -> None:
        """
        Inserta imagen de firma en Excel soportando path local (dev) y URL remota (prod/Cloudinary).
        Usa BytesIO para URLs: openpyxl 3.1+ lee todo en __init__ y no necesita el archivo al guardar.
        """
        if not image_path:
            return
        try:
            source = str(image_path).strip()
            parsed = urlparse(source)
            if parsed.scheme in ('http', 'https'):
                with urllib.request.urlopen(source) as resp:
                    img_source = io.BytesIO(resp.read())
            else:
                img_source = source
            img = Image(img_source)
            img.width = min(img.width, 180)
            img.height = min(img.height, 60)
            ws.add_image(img, anchor_cell)
        except Exception:
            return

    def _apply_formatting(self, ws: Worksheet) -> None:
        """Aplica formato a todo el worksheet."""
        for col, width in ColumnWidths.WIDTHS.items():
            ws.column_dimensions[col].width = width

    def _apply_page_setup(self, ws: Worksheet) -> None:
        """Configura la página para impresión."""
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paper_size = 9
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0 # Allow multiple pages vertically

    def _extract_nivel_escolar(self, analisis_data: Dict) -> str:
        """Extraer nombre del nivel escolar."""
        analisis_por_nivel = analisis_data.get('analisis_por_nivel', [])
        if not analisis_por_nivel:
            return "N/A"
        for nivel_data in analisis_por_nivel:
            if nivel_data.get('es_programa_actual'):
                return nivel_data.get('nivel_escolar', {}).get('nombre', 'N/A')
        return analisis_por_nivel[0].get('nivel_escolar', {}).get('nombre', 'N/A')

    def _extract_nivel_data(self, analisis_data: Dict, nivel_escolar_id: Optional[str] = None) -> Optional[Dict]:
        """
        Extraer los datos del nivel escolar específico o del nivel actual del programa.

        Args:
            analisis_data: Diccionario con todos los análisis por nivel.
            nivel_escolar_id: ID del nivel escolar específico (opcional).

        Returns:
            Diccionario con los datos del nivel escolar o None si no se encuentra.
        """
        analisis_por_nivel = analisis_data.get('analisis_por_nivel', [])

        if not analisis_por_nivel:
            return None

        # Si se especificó un nivel escolar ID, buscarlo
        if nivel_escolar_id:
            for nivel_data in analisis_por_nivel:
                nivel_id = nivel_data.get('nivel_escolar', {}).get('id', '')
                if str(nivel_id) == str(nivel_escolar_id):
                    return nivel_data

        # Si no se especificó o no se encontró, buscar el nivel del programa actual
        for nivel_data in analisis_por_nivel:
            if nivel_data.get('es_programa_actual'):
                return nivel_data

        # Como último recurso, devolver el primer nivel
        return analisis_por_nivel[0] if analisis_por_nivel else None

    def _draw_single_report(self, ws: Worksheet, start_row: int, analisis_data: Dict, nivel_escolar_id: Optional[str] = None, show_title: bool = True) -> int:
        """
        Dibuja un reporte de análisis completo en la hoja y fila especificadas.

        Args:
            ws: Hoja de Excel donde dibujar.
            start_row: Fila inicial para el reporte.
            analisis_data: Datos del análisis nutricional.
            nivel_escolar_id: ID del nivel escolar (opcional).

        Returns:
            int: La última fila utilizada por el reporte (incluyendo firmas).
        """
        menu_info = analisis_data.get('menu', {})
        logo_ref = menu_info.get('logo_path') or menu_info.get('logo_url')
        self._insert_logo(ws, logo_ref, f'A{start_row}')

        nivel_data = self._extract_nivel_data(analisis_data, nivel_escolar_id)
        if not nivel_data:
            raise ValueError("No se encontró información del nivel escolar para el reporte")

        # Detectar si es programa sin niveles (Comedores, Adulto Mayor)
        is_comedores = nivel_data.get('nivel_escolar', {}).get('id') == 'general'
        layout = self.comedores_layout if is_comedores else self.layout

        # --- Título (solo en el primer menú de la hoja) ---
        if show_title:
            title_text = "ANÁLISIS FISICO QUIMICO NUTRICIONAL DE MENUS" if is_comedores else "ANÁLISIS NUTRICIONAL"
            if is_comedores:
                title_col_start, title_col_end = 3, 6   # C:F
            else:
                title_col_start, title_col_end = 4, layout.TITLE_END_COL
            title_cell = ws.cell(row=start_row, column=title_col_start)
            title_cell.value = title_text
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(
                start_row=start_row, start_column=title_col_start,
                end_row=start_row, end_column=title_col_end
            )

        if is_comedores:
            # Sección admin compacta: solo "MENÚ No." en una fila
            menu_row = start_row + 1
            ws.cell(row=menu_row, column=1).value = "MENÚ No."
            ws.cell(row=menu_row, column=1).font = Font(bold=True)
            ws.cell(row=menu_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=menu_row, column=4).value = str(menu_info.get('nombre', 'N/A')).upper()
            ws.cell(row=menu_row, column=4).alignment = Alignment(horizontal='left', vertical='center')
            # Sin fila en blanco entre MENÚ No. y encabezados
            header_start_row = start_row + 2
        else:
            # Sección administrativa completa (7 filas)
            admin_start_row = start_row + 1
            nivel_escolar_nombre = nivel_data.get('nivel_escolar', {}).get('nombre', 'N/A')
            self._populate_administrative_section(
                ws,
                start_row=admin_start_row,
                analisis_data=analisis_data,
                nivel_escolar_nombre=nivel_escolar_nombre
            )
            for i in range(7):
                r = admin_start_row + i
                ws.merge_cells(f'A{r}:C{r}')
            header_start_row = start_row + 9

        self._populate_headers(ws, start_row=header_start_row, is_comedores=is_comedores)

        data_start_row = header_start_row + 1
        last_data_row = self._populate_ingredients_from_analysis(
            ws, start_row=data_start_row, nivel_data=nivel_data, is_comedores=is_comedores
        )

        self._add_calculations_section(ws, last_data_row, nivel_data, is_comedores=is_comedores)

        if is_comedores:
            # Sin sección de firmas para Comedores / Adulto Mayor
            # total_row=+2, req_row=+3, pct_row=+4 → dejar 2 filas de margen
            final_row = last_data_row + 6
        else:
            signatures_start_row = last_data_row + 6
            self._add_signatures(ws, signatures_start_row, menu_info=menu_info)
            # Sección de firmas ocupa aproximadamente 5-6 filas
            final_row = signatures_start_row + 5

        return final_row
