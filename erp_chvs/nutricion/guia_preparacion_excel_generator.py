"""
Generador de Excel: Guias de preparacion.

Estructura funcional equivalente al formato solicitado:
- Una pestana por menu.
- Bloque administrativo por hoja.
- Tabla por preparacion con filas de ingredientes.
- Columnas por nivel: peso bruto, peso neto y peso servido.
"""

import io
import re
import unicodedata
import urllib.request
from decimal import Decimal
from typing import Dict, List, Optional, Tuple
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from principal.models import TablaGradosEscolaresUapa

from fuzzywuzzy import fuzz

from .models import (
    FirmaNutricionalContrato,
    ProcedimientoPreparacion,
    TablaAnalisisNutricionalMenu,
    TablaIngredientesPorNivel,
    TablaMenus,
    TablaPreparacionIngredientes,
    TablaPreparaciones,
)

UMBRAL_COINCIDENCIA = 70  # porcentaje mínimo para aceptar un match


NIVELES_ORDEN_PAE = [
    "prescolar",
    "primaria_1_2_3",
    "primaria_4_5",
    "secundaria",
    "media_ciclo_complementario",
]

# Alias de compatibilidad (no eliminar — puede ser importado externamente)
NIVELES_ORDEN = NIVELES_ORDEN_PAE

NIVELES_LABEL = {
    "prescolar": "PREESCOLAR",
    "primaria_1_2_3": "PRIMARIA: 1, 2 Y 3",
    "primaria_4_5": "PRIMARIA: 4 Y 5",
    "secundaria": "SECUNDARIA",
    "media_ciclo_complementario": "MEDIA Y CICLO COMPLEMENTARIO",
    "general": "COMEDORES COMUNITARIOS",
}


class GuiaPreparacionExcelGenerator:
    def __init__(self):
        side = Side(style="thin", color="000000")
        self.border = Border(left=side, right=side, top=side, bottom=side)
        self.header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        self.center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # Configuración dinámica de niveles; sobreescrita en generate()
        self.niveles_orden: List[str] = list(NIVELES_ORDEN_PAE)
        self.col_proc: int = 18  # columna de PROCEDIMIENTO (PAE default)

    def _cargar_catalogo(self) -> List[Tuple[str, str]]:
        """Carga el catálogo activo de procedimientos: [(nombre_normalizado, procedimiento)]."""
        return [
            (self._normalizar(p.nombre), p.procedimiento)
            for p in ProcedimientoPreparacion.objects.filter(activo=True)
        ]

    @staticmethod
    def _normalizar(texto: str) -> str:
        texto = unicodedata.normalize("NFKD", str(texto or ""))
        texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
        return texto.lower().strip()

    def _buscar_procedimiento(self, nombre_prep: str, catalogo: List[Tuple[str, str]]) -> str:
        """Retorna el procedimiento con mayor coincidencia, o '' si no supera el umbral."""
        if not catalogo:
            return ""
        nombre_norm = self._normalizar(nombre_prep)
        mejor_score = 0
        mejor_texto = ""
        for nombre_cat, procedimiento in catalogo:
            score = fuzz.token_sort_ratio(nombre_norm, nombre_cat)
            if score > mejor_score:
                mejor_score = score
                mejor_texto = procedimiento
        return mejor_texto if mejor_score >= UMBRAL_COINCIDENCIA else ""

    def generate(self, programa_id: int, modalidad_id: str) -> io.BytesIO:
        # Determinar si el programa usa niveles educativos (PAE) o es un programa sin niveles
        # (Comedores Comunitarios, Adulto Mayor, etc.)
        from planeacion.models import Programa as ProgramaModel
        try:
            programa_obj = ProgramaModel.objects.select_related('tipo_programa').get(id=programa_id)
            tiene_niveles = getattr(getattr(programa_obj, 'tipo_programa', None), 'tiene_niveles', True)
        except ProgramaModel.DoesNotExist:
            tiene_niveles = True

        self.niveles_orden = list(NIVELES_ORDEN_PAE) if tiene_niveles else ['general']
        self.col_proc = 2 + len(self.niveles_orden) * 3 + 1

        menus = list(
            TablaMenus.objects.filter(
                id_contrato_id=programa_id,
                id_modalidad_id=modalidad_id,
            )
            .select_related("id_contrato", "id_contrato__tipo_programa", "id_modalidad")
        )
        menus.sort(key=self._menu_sort_key)

        wb = Workbook()
        wb.remove(wb.active)

        if not menus:
            ws = wb.create_sheet("SIN DATOS")
            ws["A1"] = "No se encontraron menus para el programa/modalidad seleccionados."
            stream = io.BytesIO()
            wb.save(stream)
            stream.seek(0)
            return stream

        catalogo = self._cargar_catalogo()

        for menu in menus:
            ws = wb.create_sheet(title=f"Menu {menu.menu}"[:31])
            self._build_sheet(ws, menu, catalogo)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    def _build_sheet(self, ws, menu: TablaMenus, catalogo: List[Tuple[str, str]]) -> None:
        self._set_columns(ws)
        self._draw_top(ws, menu)
        row = 9
        row = self._draw_table_header(ws, row)
        row_end = self._draw_table_body(ws, row, menu, catalogo)
        self._draw_signature_block(ws, row_end, menu)
        self._apply_all_borders(ws)

    @staticmethod
    def _menu_sort_key(menu: TablaMenus):
        menu_text = str(menu.menu).strip()
        match = re.search(r"\d+", menu_text)
        if match:
            return (0, int(match.group()), menu_text.lower())
        return (1, float("inf"), menu_text.lower())

    def _set_columns(self, ws):
        ws.column_dimensions["A"].width = 24  # preparacion
        ws.column_dimensions["B"].width = 22  # ingrediente
        for col_idx in range(3, self.col_proc):
            ws.column_dimensions[get_column_letter(col_idx)].width = 9
        ws.column_dimensions[get_column_letter(self.col_proc)].width = 42  # procedimiento

    def _draw_top(self, ws, menu: TablaMenus):
        lc = get_column_letter(self.col_proc)  # última columna (ej: "R" para PAE, "F" para Comedores)
        ws.row_dimensions[1].height = 38
        ws.row_dimensions[2].height = 38
        ws.merge_cells(f"A1:{lc}2")
        ws["A1"].fill = self.header_fill
        ws["A1"].border = self.border

        # Logo programa (si existe)
        if menu.id_contrato and menu.id_contrato.imagen:
            # Anclar logo en la penúltima columna de la fila 1
            logo_col = get_column_letter(max(self.col_proc - 3, 2))
            self._insert_image_field(ws, menu.id_contrato.imagen, f"{logo_col}1", max_w=180, max_h=60)

        ws.merge_cells(f"A3:{lc}3")
        ws["A3"] = "GUIA DE PREPARACION DE ALIMENTOS Y ESTANDARIZACION DE RECETAS"
        ws["A3"].font = Font(bold=True, size=11)
        ws["A3"].alignment = self.center
        ws["A3"].fill = self.header_fill
        ws["A3"].border = self.border

        ws.merge_cells(f"A4:{lc}4")
        ws["A4"] = "COMPLEMENTO ALIMENTARIO JORNADA AM/PM PREPARADO EN SITIO"
        ws["A4"].font = Font(bold=True, size=11)
        ws["A4"].alignment = self.center
        ws["A4"].fill = self.header_fill
        ws["A4"].border = self.border

        if len(self.niveles_orden) > 1:
            # --- Layout PAE: filas 5-7 con sub-divisiones por escolaridad y grupo étnico ---
            ws.merge_cells("B5:D5")
            ws["B5"] = "PREESCOLAR - MEDIA Y CICLO COMPLEMENTARIO"
            ws["B5"].alignment = self.center
            ws["B5"].font = Font(bold=True)
            ws["B5"].fill = self.header_fill
            ws["B5"].border = self.border

            ws.merge_cells("E5:G5")
            ws["E5"] = "OPERADOR"
            ws["E5"].alignment = self.center
            ws["E5"].font = Font(bold=True)
            ws["E5"].fill = self.header_fill
            ws["E5"].border = self.border

            ws.merge_cells(f"H5:{lc}5")
            ws["H5"] = str(menu.id_contrato.programa).upper()
            ws["H5"].alignment = self.center
            ws["H5"].font = Font(bold=True)
            ws["H5"].fill = self.header_fill
            ws["H5"].border = self.border

            # Fila 6
            ws["A6"] = "ESCOLARIDAD"
            ws["A6"].font = Font(bold=True)
            ws["A6"].alignment = self.center
            ws["A6"].fill = self.header_fill
            ws["A6"].border = self.border

            ws["B6"] = "INDÍGENA"
            ws["B6"].alignment = self.center
            ws["B6"].fill = self.header_fill
            ws["B6"].border = self.border

            ws.merge_cells("D6:E6")
            ws["D6"] = "COMUNIDAD / PUEBLO INDÍGENA"
            ws["D6"].alignment = self.center
            ws["D6"].fill = self.header_fill
            ws["D6"].border = self.border

            ws.merge_cells("F6:G6")
            ws["F6"] = ""
            ws["F6"].alignment = self.center
            ws["F6"].fill = self.header_fill
            ws["F6"].border = self.border

            ws.merge_cells("H6:I6")
            ws["H6"] = "AFROCOLOMBIANO / PALENQUEROS"
            ws["H6"].alignment = self.center
            ws["H6"].fill = self.header_fill
            ws["H6"].border = self.border

            if self.col_proc >= 10:
                ws.merge_cells(f"J6:{lc}6")
                ws["J6"] = ""
                ws["J6"].alignment = self.center
                ws["J6"].fill = self.header_fill
                ws["J6"].border = self.border

            # Fila 7
            ws["A7"] = "GRUPO ETNICO"
            ws["A7"].alignment = self.center
            ws["A7"].font = Font(bold=True)
            ws["A7"].fill = self.header_fill
            ws["A7"].border = self.border

            ws["B7"] = "RAIZAL"
            ws["B7"].alignment = self.center
            ws["B7"].fill = self.header_fill
            ws["B7"].border = self.border

            ws.merge_cells("D7:E7")
            ws["D7"] = "ROM"
            ws["D7"].alignment = self.center
            ws["D7"].fill = self.header_fill
            ws["D7"].border = self.border

            ws.merge_cells("F7:G7")
            ws["F7"] = ""
            ws["F7"].alignment = self.center
            ws["F7"].fill = self.header_fill
            ws["F7"].border = self.border

            ws.merge_cells("H7:I7")
            ws["H7"] = "SIN PERTENENCIA ÉTNICA"
            ws["H7"].alignment = self.center
            ws["H7"].fill = self.header_fill
            ws["H7"].border = self.border

            if self.col_proc >= 10:
                ws.merge_cells(f"J7:{lc}7")
                ws["J7"] = "X"
                ws["J7"].alignment = self.center
                ws["J7"].font = Font(bold=True, size=13)
                ws["J7"].fill = self.header_fill
                ws["J7"].border = self.border

        else:
            # --- Layout Comedores: filas 5-7 simplificadas ---
            ws.merge_cells(f"A5:{lc}5")
            ws["A5"] = str(menu.id_contrato.programa).upper() if menu.id_contrato else ""
            ws["A5"].alignment = self.center
            ws["A5"].font = Font(bold=True)
            ws["A5"].fill = self.header_fill
            ws["A5"].border = self.border

            ws.merge_cells(f"A6:B6")
            ws["A6"] = "GRUPO ETNICO"
            ws["A6"].alignment = self.center
            ws["A6"].font = Font(bold=True)
            ws["A6"].fill = self.header_fill
            ws["A6"].border = self.border

            ws.merge_cells(f"C6:{lc}6")
            ws["C6"] = "SIN PERTENENCIA ÉTNICA"
            ws["C6"].alignment = self.center
            ws["C6"].fill = self.header_fill
            ws["C6"].border = self.border

            ws.merge_cells(f"A7:{lc}7")
            ws["A7"] = ""
            ws["A7"].fill = self.header_fill
            ws["A7"].border = self.border

        ws.merge_cells(f"A8:{lc}8")
        ws["A8"] = f"Menu {menu.menu}"
        ws["A8"].font = Font(bold=True)
        ws["A8"].alignment = self.center
        ws["A8"].fill = self.header_fill
        ws["A8"].border = self.border

    def _draw_table_header(self, ws, start_row: int) -> int:
        row1 = start_row
        row2 = start_row + 1

        ws.merge_cells(start_row=row1, start_column=1, end_row=row2, end_column=1)
        ws.cell(row=row1, column=1, value="PREPARACION")
        ws.cell(row=row1, column=1).font = Font(bold=True)
        ws.cell(row=row1, column=1).alignment = self.center
        ws.cell(row=row1, column=1).fill = self.header_fill

        ws.merge_cells(start_row=row1, start_column=2, end_row=row2, end_column=2)
        ws.cell(row=row1, column=2, value="NOMBRE DEL ALIMENTO\n(Ingredientes)")
        ws.cell(row=row1, column=2).font = Font(bold=True)
        ws.cell(row=row1, column=2).alignment = self.center
        ws.cell(row=row1, column=2).fill = self.header_fill

        col = 3
        for nivel_id in self.niveles_orden:
            ws.merge_cells(start_row=row1, start_column=col, end_row=row1, end_column=col + 2)
            ws.cell(row=row1, column=col, value=NIVELES_LABEL.get(nivel_id, nivel_id.upper()))
            ws.cell(row=row1, column=col).font = Font(bold=True)
            ws.cell(row=row1, column=col).alignment = self.center
            ws.cell(row=row1, column=col).fill = self.header_fill

            ws.cell(row=row2, column=col, value="PESO\nBRUTO (g)")
            ws.cell(row=row2, column=col + 1, value="PESO\nNETO (g)")
            ws.cell(row=row2, column=col + 2, value="PESO\nSERVIDO (g)")
            for c in range(col, col + 3):
                ws.cell(row=row2, column=c).font = Font(bold=True, size=9)
                ws.cell(row=row2, column=c).alignment = self.center
                ws.cell(row=row2, column=c).fill = self.header_fill
            col += 3

        ws.merge_cells(start_row=row1, start_column=self.col_proc, end_row=row2, end_column=self.col_proc)
        ws.cell(row=row1, column=self.col_proc, value="PROCEDIMIENTO DE PREPARACION")
        ws.cell(row=row1, column=self.col_proc).font = Font(bold=True)
        ws.cell(row=row1, column=self.col_proc).alignment = self.center
        ws.cell(row=row1, column=self.col_proc).fill = self.header_fill

        for r in (row1, row2):
            for c in range(1, self.col_proc + 1):
                ws.cell(row=r, column=c).border = self.border

        return row2 + 1

    # Componentes que representan bebida/sobremesa (porción servida fija = 200 ml/g)
    _COMPONENTES_BEBIDA = {'com1', 'com14'}
    # Modalidades industrializadas: no aplica la porción fija de 200g
    _MODALIDADES_INDUSTRIALIZADAS = {'020511', '20502'}

    def _draw_table_body(self, ws, start_row: int, menu: TablaMenus, catalogo: List[Tuple[str, str]]) -> int:
        niveles_por_columna = self._get_niveles_por_columna()
        modalidad_id = str(menu.id_modalidad_id or '').strip()
        es_industrializado = modalidad_id in self._MODALIDADES_INDUSTRIALIZADAS

        from nutricion.utils.orden_componentes import sort_preparaciones_objetos
        preps_qs = list(
            TablaPreparaciones.objects.filter(id_menu=menu).select_related('id_componente')
        )
        preps = sort_preparaciones_objetos(preps_qs, menu.id_modalidad_id)
        rels = list(
            TablaPreparacionIngredientes.objects.filter(id_preparacion__in=preps)
            .select_related("id_preparacion", "id_ingrediente_siesa")
            .order_by("id_preparacion__preparacion", "id_ingrediente_siesa__nombre_del_alimento")
        )

        analisis_por_nivel = {
            a.id_nivel_escolar_uapa_id: a
            for a in TablaAnalisisNutricionalMenu.objects.filter(id_menu=menu)
        }
        ings_guardados = list(
            TablaIngredientesPorNivel.objects.filter(
                id_analisis__in=list(analisis_por_nivel.values()),
                id_preparacion__in=preps,
            )
        )
        idx = self._build_index(ings_guardados)

        rels_by_prep: Dict[int, List[TablaPreparacionIngredientes]] = {}
        for rel in rels:
            rels_by_prep.setdefault(rel.id_preparacion_id, []).append(rel)

        row = start_row
        for prep in preps:
            prep_rels = rels_by_prep.get(prep.id_preparacion, [])
            if not prep_rels:
                continue

            # Procedimiento: buscar en catálogo por fuzzy matching del nombre
            procedimiento_texto = self._buscar_procedimiento(prep.preparacion, catalogo)

            # Bebidas (com1/com14) en modalidades no industrializadas → porción fija 200g
            es_bebida = str(prep.id_componente_id or '') in self._COMPONENTES_BEBIDA
            porcion_fija_bebida = es_bebida and not es_industrializado

            # Peso servido por nivel = 200g (bebida) ó Σ(neto × FC) por ingrediente
            peso_servido_by_nivel = {}
            for _, nivel_id in niveles_por_columna:
                if not nivel_id:
                    continue
                if porcion_fija_bebida:
                    peso_servido_by_nivel[nivel_id] = Decimal("200")
                else:
                    total = Decimal("0")
                    for rel in prep_rels:
                        _, neto = self._get_bruto_neto(rel, prep.id_preparacion, nivel_id, idx, analisis_por_nivel)
                        fc = Decimal(str(rel.id_ingrediente_siesa.factor_coccion or 1))
                        total += neto * fc
                    peso_servido_by_nivel[nivel_id] = total

            prep_start = row
            for rel in prep_rels:
                ws.cell(row=row, column=2, value=rel.id_ingrediente_siesa.nombre_del_alimento)
                ws.cell(row=row, column=2).alignment = self.left
                ws.cell(row=row, column=2).border = self.border

                col = 3
                for _, nivel_id in niveles_por_columna:
                    bruto, neto = self._get_bruto_neto(rel, prep.id_preparacion, nivel_id, idx, analisis_por_nivel)

                    ws.cell(row=row, column=col, value=float(round(bruto, 2)))
                    ws.cell(row=row, column=col + 1, value=float(round(neto, 2)))
                    ws.cell(row=row, column=col + 2, value="")
                    for c in (col, col + 1, col + 2):
                        ws.cell(row=row, column=c).alignment = self.center
                        ws.cell(row=row, column=c).border = self.border
                    col += 3

                ws.cell(row=row, column=self.col_proc, value="")
                ws.cell(row=row, column=self.col_proc).alignment = self.left
                ws.cell(row=row, column=self.col_proc).border = self.border
                row += 1

            prep_end = row - 1
            ws.merge_cells(start_row=prep_start, start_column=1, end_row=prep_end, end_column=1)
            ws.cell(row=prep_start, column=1, value=prep.preparacion.upper())
            ws.cell(row=prep_start, column=1).font = Font(bold=True)
            ws.cell(row=prep_start, column=1).alignment = self.center
            ws.cell(row=prep_start, column=1).border = self.border

            # Peso servido combinado por preparacion (una sola celda vertical por nivel)
            col = 3
            for _, nivel_id in niveles_por_columna:
                servido = peso_servido_by_nivel.get(nivel_id, Decimal("0")) if nivel_id else Decimal("0")
                servido_col = col + 2
                if prep_end > prep_start:
                    ws.merge_cells(
                        start_row=prep_start,
                        start_column=servido_col,
                        end_row=prep_end,
                        end_column=servido_col,
                    )
                ws.cell(row=prep_start, column=servido_col, value=float(round(servido, 2)))
                ws.cell(row=prep_start, column=servido_col).alignment = self.center
                ws.cell(row=prep_start, column=servido_col).border = self.border
                col += 3

            # Procedimiento: texto del catálogo (o en blanco si no hubo match)
            ws.merge_cells(start_row=prep_start, start_column=self.col_proc, end_row=prep_end, end_column=self.col_proc)
            ws.cell(row=prep_start, column=self.col_proc, value=procedimiento_texto)
            ws.cell(row=prep_start, column=self.col_proc).alignment = self.left
            ws.cell(row=prep_start, column=self.col_proc).border = self.border

            # borde columna A y columna procedimiento en filas internas de merge
            for r in range(prep_start, prep_end + 1):
                ws.cell(row=r, column=1).border = self.border
                ws.cell(row=r, column=self.col_proc).border = self.border

        ws.freeze_panes = "C11"
        return row

    def _build_index(self, ingredientes_guardados: List[TablaIngredientesPorNivel]) -> Dict[Tuple[str, int, str], TablaIngredientesPorNivel]:
        idx = {}
        for item in ingredientes_guardados:
            nivel_id = item.id_analisis.id_nivel_escolar_uapa_id
            prep_id = item.id_preparacion_id
            codigo = str(item.codigo_icbf or item.id_ingrediente_siesa_id or "").strip()
            if not codigo:
                continue
            idx[(nivel_id, prep_id, codigo)] = item
        return idx

    @staticmethod
    def _normalize_nivel_key(value: str) -> str:
        value = str(value or "").strip().lower()
        value = unicodedata.normalize("NFKD", value)
        value = "".join(ch for ch in value if not unicodedata.combining(ch))
        value = value.replace("-", "_").replace(" ", "_")
        return value

    def _get_niveles_por_columna(self) -> List[Tuple[str, Optional[str]]]:
        niveles = list(
            TablaGradosEscolaresUapa.objects.values_list(
                "id_grado_escolar_uapa",
                "nivel_escolar_uapa",
            )
        )
        por_key: Dict[str, str] = {}
        for nivel_id, nombre in niveles:
            nivel_id_str = str(nivel_id).strip()
            for candidato in (
                self._normalize_nivel_key(nivel_id_str),
                self._normalize_nivel_key(nombre),
            ):
                if candidato and candidato not in por_key:
                    por_key[candidato] = nivel_id_str

        return [(slot, por_key.get(self._normalize_nivel_key(slot))) for slot in self.niveles_orden]

    def _get_bruto_neto(
        self,
        rel: TablaPreparacionIngredientes,
        prep_id: int,
        nivel_id: Optional[str],
        idx: Dict[Tuple[str, int, str], TablaIngredientesPorNivel],
        analisis_por_nivel: Dict[str, TablaAnalisisNutricionalMenu],
    ) -> Tuple[Decimal, Decimal]:
        guardado = None
        if nivel_id:
            codigo = str(rel.id_ingrediente_siesa.codigo).strip()
            guardado = idx.get((nivel_id, prep_id, codigo))

        if guardado:
            neto = Decimal(guardado.peso_neto or 0)
        else:
            neto = Decimal(rel.gramaje or 0)

        parte_comestible = Decimal(rel.id_ingrediente_siesa.parte_comestible_field or 100)
        if parte_comestible <= 0:
            parte_comestible = Decimal("100")
        bruto = (neto * Decimal("100")) / parte_comestible
        return bruto, neto

    def _draw_signature_block(self, ws, start_row: int, menu: TablaMenus) -> None:
        """
        Bloque inferior de firmas usando nutricion_firma_nutricional_contrato.
        """
        ws.row_dimensions[start_row].height = 50
        ws.row_dimensions[start_row + 1].height = 50
        firma_cfg = FirmaNutricionalContrato.objects.filter(programa=menu.id_contrato).first()
        elabora_nombre = (firma_cfg.elabora_nombre if firma_cfg else "") or ""
        elabora_matricula = (firma_cfg.elabora_matricula if firma_cfg else "") or ""
        elabora_firma_texto = (firma_cfg.elabora_firma_texto if firma_cfg else "") or ""
        aprueba_nombre = (firma_cfg.aprueba_nombre if firma_cfg else "") or ""
        aprueba_matricula = (firma_cfg.aprueba_matricula if firma_cfg else "") or ""
        aprueba_firma_texto = (firma_cfg.aprueba_firma_texto if firma_cfg else "") or ""

        elabora_img = None
        aprueba_img = None
        if firma_cfg:
            elabora_img = self._get_image_source(firma_cfg.elabora_firma_imagen)
            aprueba_img = self._get_image_source(firma_cfg.aprueba_firma_imagen)

        rows = [start_row, start_row + 1]
        for r in rows:
            for c in range(1, self.col_proc + 1):
                ws.cell(row=r, column=c).border = self.border
                ws.cell(row=r, column=c).alignment = self.center
                ws.cell(row=r, column=c).fill = self.header_fill

        # Dividir el espacio disponible en 4 zonas: etiqueta, nombre, matrícula label, matrícula valor+firma
        # Para PAE (18 cols): etiqueta=1-4, nombre=5-8, mat_label=9-11, mat_val=12-13, firma_label=14, firma=15-18
        # Para Comedores (6 cols): se ajusta proporcionalmente
        cp = self.col_proc
        # Zona 1 (etiqueta): cols 1-max(1, cp//4)
        z1_end = max(1, cp // 4)
        # Zona 2 (nombre): cols z1_end+1 - z1_end+max(1,(cp-z1_end)//3)
        z2_end = min(z1_end + max(1, (cp - z1_end) // 3), cp - 1)
        # Zona firma: cols z2_end+1 - cp
        firma_col = z2_end + 1
        firma_anchor = get_column_letter(firma_col)

        # Fila 1 (elabora)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=z1_end)
        ws.cell(row=start_row, column=1, value="NOMBRE NUTRICIONISTA - DIETISTA POR PARTE DEL OPERADOR")
        ws.cell(row=start_row, column=1).alignment = self.left
        ws.cell(row=start_row, column=1).font = Font(bold=True)

        ws.merge_cells(start_row=start_row, start_column=z1_end + 1, end_row=start_row, end_column=z2_end)
        ws.cell(row=start_row, column=z1_end + 1, value=elabora_nombre)
        ws.cell(row=start_row, column=z1_end + 1).font = Font(bold=True)

        ws.merge_cells(start_row=start_row, start_column=firma_col, end_row=start_row, end_column=cp)
        if elabora_img:
            self._insert_signature_image(ws, elabora_img, f"{firma_anchor}{start_row}")
        else:
            ws.cell(row=start_row, column=firma_col, value=elabora_firma_texto)
            ws.cell(row=start_row, column=firma_col).alignment = self.left

        # Fila 2 (aprueba)
        ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=z1_end)
        ws.cell(
            row=start_row + 1,
            column=1,
            value="NOMBRE NUTRICIONISTA - DIETISTA QUE APRUEBA LA GUIA POR PARTE DEL PAE SEM"
        )
        ws.cell(row=start_row + 1, column=1).alignment = self.left
        ws.cell(row=start_row + 1, column=1).font = Font(bold=True)

        ws.merge_cells(start_row=start_row + 1, start_column=z1_end + 1, end_row=start_row + 1, end_column=z2_end)
        ws.cell(row=start_row + 1, column=z1_end + 1, value=aprueba_nombre)
        ws.cell(row=start_row + 1, column=z1_end + 1).font = Font(bold=True)

        ws.merge_cells(start_row=start_row + 1, start_column=firma_col, end_row=start_row + 1, end_column=cp)
        if aprueba_img:
            self._insert_signature_image(ws, aprueba_img, f"{firma_anchor}{start_row + 1}")
        else:
            ws.cell(row=start_row + 1, column=firma_col, value=aprueba_firma_texto)
            ws.cell(row=start_row + 1, column=firma_col).alignment = self.left

    @staticmethod
    def _get_image_source(field) -> str | None:
        """Retorna path (dev/local) o URL (prod/Cloudinary) de un FieldFile. None si no hay imagen."""
        if not field:
            return None
        try:
            return field.path
        except (FileNotFoundError, ValueError, NotImplementedError):
            pass
        try:
            return field.url
        except (ValueError, NotImplementedError):
            return None

    def _insert_image_field(self, ws, field, anchor_cell: str, max_w: int = 200, max_h: int = 60) -> None:
        """Inserta imagen desde un FieldFile Django en una celda, soportando local y Cloudinary."""
        source = self._get_image_source(field)
        self._insert_signature_image(ws, source, anchor_cell, max_w=max_w, max_h=max_h)

    @staticmethod
    def _insert_signature_image(ws, image_source: str | None, anchor_cell: str, max_w: int = 240, max_h: int = 70) -> None:
        """Inserta imagen en Excel desde path local o URL remota (Cloudinary).
        Usa BytesIO para URLs: openpyxl 3.1+ lee todo en __init__ y no necesita el archivo al guardar."""
        if not image_source:
            return
        try:
            source = str(image_source).strip()
            parsed = urlparse(source)
            if parsed.scheme in ('http', 'https'):
                with urllib.request.urlopen(source) as resp:
                    img_source = io.BytesIO(resp.read())
            else:
                img_source = source
            img = Image(img_source)
            img.width = min(img.width, max_w)
            img.height = min(img.height, max_h)
            ws.add_image(img, anchor_cell)
        except Exception:
            return

    def _apply_all_borders(self, ws) -> None:
        max_row = ws.max_row
        for row in range(1, max_row + 1):
            for col in range(1, self.col_proc + 1):
                ws.cell(row=row, column=col).border = self.border
