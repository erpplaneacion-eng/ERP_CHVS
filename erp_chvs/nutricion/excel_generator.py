"""
Generador de análisis nutricional en formato Excel.

Este módulo genera archivos Excel con análisis nutricional completo
para un solo menú.
"""

import io
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from .services.analisis_service import AnalisisNutricionalService
from .excel_drawing_utils import ExcelReportDrawer


class NutritionalAnalysisExcelGenerator(ExcelReportDrawer):
    """
    Generador de análisis nutricional en Excel para un solo menú.
    Hereda la lógica de dibujo de ExcelReportDrawer.
    """

    def __init__(self):
        super().__init__()

    def generate_from_analysis_service(
        self,
        menu_id: int,
        nivel_escolar_id: Optional[str] = None
    ) -> io.BytesIO:
        """
        Genera un WORKBOOK completo para un solo menú, con una pestaña por
        cada nivel escolar.
        """
        try:
            analisis_data = AnalisisNutricionalService.obtener_analisis_completo(menu_id)

            if not analisis_data.get('success'):
                raise ValueError("No se pudo obtener el análisis del menú")

            wb = Workbook()
            wb.remove(wb.active)  # Eliminar la hoja por defecto

            analysis_by_level = analisis_data.get("analisis_por_nivel", [])

            # Si se pide un nivel específico, filtrar solo ese
            if nivel_escolar_id:
                analysis_by_level = [
                    nivel for nivel in analysis_by_level 
                    if str(nivel.get('nivel_escolar', {}).get('id')) == str(nivel_escolar_id)
                ]

            if not analysis_by_level:
                raise ValueError("No se encontraron datos de análisis para los niveles especificados.")

            for nivel_data in analysis_by_level:
                nivel_nombre = nivel_data.get('nivel_escolar', {}).get('nombre', 'Nivel')
                sheet_name = nivel_nombre[:31]
                ws = wb.create_sheet(title=sheet_name)

                current_nivel_id = nivel_data.get('nivel_escolar', {}).get('id')

                self._draw_single_report(ws, 1, analisis_data, nivel_escolar_id=current_nivel_id)

                self._apply_formatting(ws)
                self._apply_page_setup(ws)

            stream = io.BytesIO()
            wb.save(stream)
            stream.seek(0)
            return stream

        except Exception as e:
            return self._generate_error_excel(f"Error generando Excel: {str(e)}")

    def _generate_error_excel(self, error_message: str) -> io.BytesIO:
        """Generar Excel con mensaje de error."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Error"
        ws['A1'] = f"Error: {error_message}"
        ws['A1'].font = Font(bold=True, color="FF0000")
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

# =====================================================================
# FUNCIONES DE COMPATIBILIDAD (WRAPPER)
# =====================================================================

def generate_excel_from_service(
    menu_id: int,
    nivel_escolar_id: Optional[str] = None
) -> io.BytesIO:
    """
    Genera Excel usando el servicio de análisis nutricional.
    Función de compatibilidad.
    """
    generator = NutritionalAnalysisExcelGenerator()
    return generator.generate_from_analysis_service(menu_id, nivel_escolar_id)

def generate_menu_excel(menu_id: int, nivel_escolar_id: Optional[str] = None) -> io.BytesIO:
    """Genera análisis nutricional completo."""
    generator = NutritionalAnalysisExcelGenerator()
    return generator.generate_from_analysis_service(menu_id, nivel_escolar_id)

def generate_menu_excel_with_nivel(menu_id: int, nivel_escolar_id: str) -> io.BytesIO:
    """Genera análisis nutricional para un nivel escolar específico."""
    generator = NutritionalAnalysisExcelGenerator()
    return generator.generate_from_analysis_service(menu_id, nivel_escolar_id)

def generate_advanced_nutritional_excel(
    menu_id: int,
    nivel_escolar_id: Optional[str] = None,
    use_saved_analysis: bool = True
) -> io.BytesIO:
    """Genera Excel nutricional avanzado (función de compatibilidad)."""
    generator = NutritionalAnalysisExcelGenerator()
    return generator.generate_from_analysis_service(menu_id, nivel_escolar_id)

def generate_menu_excel_real_data(menu_id: int) -> io.BytesIO:
    """
    Genera análisis nutricional usando datos reales.
    Función de compatibilidad - usa el servicio de análisis.
    """
    generator = NutritionalAnalysisExcelGenerator()
    return generator.generate_from_analysis_service(menu_id)