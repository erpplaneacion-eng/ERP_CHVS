import json
from io import BytesIO
from datetime import date
from decimal import Decimal
from pathlib import Path

from django.contrib.auth.models import User
from django.test import Client, TestCase
from django.urls import reverse

from planeacion.models import Programa
from principal.models import ModalidadesDeConsumo, PrincipalMunicipio, TablaGradosEscolaresUapa
from openpyxl import load_workbook

from .excel_generator import generate_menu_excel
from .excel_drawing_utils import ExcelReportDrawer
from .models import (
    ComponentesAlimentos,
    FirmaNutricionalContrato,
    GruposAlimentos,
    MinutaPatronMeta,
    TablaAlimentos2018Icbf,
    TablaAnalisisNutricionalMenu,
    TablaIngredientesPorNivel,
    TablaIngredientesSiesa,
    TablaMenus,
    TablaPreparacionIngredientes,
    TablaPreparaciones,
    TablaRequerimientosNutricionales,
)
from .services.analisis_service import AnalisisNutricionalService


class Paso2PreparacionesEditorIntegrationTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_user(username="paso2_user", password="testpass123")
        cls.client = Client()

        cls.nivel_prescolar = TablaGradosEscolaresUapa.objects.create(
            id_grado_escolar_uapa="prescolar",
            nivel_escolar_uapa="Preescolar",
        )
        cls.nivel_primaria = TablaGradosEscolaresUapa.objects.create(
            id_grado_escolar_uapa="primaria",
            nivel_escolar_uapa="Primaria 1-3",
        )

        cls.modalidad = ModalidadesDeConsumo.objects.create(
            id_modalidades="modpaso2",
            modalidad="CAJM AM",
            cod_modalidad="CAJM",
        )
        cls.municipio = PrincipalMunicipio.objects.create(
            codigo_municipio=11111,
            nombre_municipio="Municipio Paso2",
            codigo_departamento="11",
        )
        cls.programa = Programa.objects.create(
            programa="Programa Paso2",
            contrato="CT-P2-001",
            municipio=cls.municipio,
            fecha_inicial=date(2025, 1, 1),
            fecha_final=date(2025, 12, 31),
            estado="activo",
        )
        cls.menu = TablaMenus.objects.create(
            menu="Menu Paso2",
            id_modalidad=cls.modalidad,
            id_contrato=cls.programa,
        )

        cls.grupo = GruposAlimentos.objects.create(
            id_grupo_alimentos="grp_p2",
            grupo_alimentos="Lacteos",
        )
        cls.componente = ComponentesAlimentos.objects.create(
            id_componente="comp_p2",
            componente="Bebida con leche",
            id_grupo_alimentos=cls.grupo,
        )

        cls.alimento_icbf = TablaAlimentos2018Icbf.objects.create(
            codigo="A001",
            nombre_del_alimento="Leche entera",
            humedad_g=Decimal("87.00"),
            energia_kcal=60,
            energia_kj=251,
            proteina_g=Decimal("3.20"),
            lipidos_g=Decimal("3.30"),
            carbohidratos_totales_g=Decimal("4.80"),
            calcio_mg=120,
            hierro_mg=Decimal("0.10"),
            sodio_mg=50,
            parte_comestible_field=100,
            id_componente=cls.componente,
        )
        cls.ingrediente_siesa = TablaIngredientesSiesa.objects.create(
            id_ingrediente_siesa="A001",
            nombre_ingrediente="Leche entera",
        )

        cls.preparacion = TablaPreparaciones.objects.create(
            preparacion="Leche preparada",
            id_menu=cls.menu,
            id_componente=cls.componente,
        )
        TablaPreparacionIngredientes.objects.create(
            id_preparacion=cls.preparacion,
            id_ingrediente_siesa=cls.alimento_icbf,
            gramaje=Decimal("100.00"),
        )

        MinutaPatronMeta.objects.create(
            id_modalidad=cls.modalidad,
            id_grado_escolar_uapa=cls.nivel_prescolar,
            id_componente=cls.componente,
            id_grupo_alimentos=cls.grupo,
            peso_neto_minimo=Decimal("120.00"),
            peso_neto_maximo=Decimal("160.00"),
        )
        MinutaPatronMeta.objects.create(
            id_modalidad=cls.modalidad,
            id_grado_escolar_uapa=cls.nivel_primaria,
            id_componente=cls.componente,
            id_grupo_alimentos=cls.grupo,
            peso_neto_minimo=Decimal("180.00"),
            peso_neto_maximo=Decimal("220.00"),
        )

        for nivel in [cls.nivel_prescolar, cls.nivel_primaria]:
            TablaRequerimientosNutricionales.objects.create(
                id_requerimiento_nutricional=f"req_{nivel.id_grado_escolar_uapa}",
                calorias_kcal=Decimal("276.0"),
                proteina_g=Decimal("9.9"),
                grasa_g=Decimal("9.6"),
                cho_g=Decimal("36.5"),
                calcio_mg=Decimal("159.0"),
                hierro_mg=Decimal("1.5"),
                sodio_mg=Decimal("95.0"),
                id_nivel_escolar_uapa=nivel,
                id_modalidad=cls.modalidad,
            )

    def setUp(self):
        self.client = Client()
        self.client.force_login(self.user)

    def test_vista_preparaciones_editor_crea_analisis_por_nivel_y_aplica_rangos_por_nivel(self):
        self.assertEqual(TablaAnalisisNutricionalMenu.objects.filter(id_menu=self.menu).count(), 0)

        url = reverse("nutricion:preparaciones_editor", args=[self.menu.id_menu])
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        niveles_data = response.context["niveles_data"]
        self.assertEqual(len(niveles_data), 2)

        analisis_count = TablaAnalisisNutricionalMenu.objects.filter(id_menu=self.menu).count()
        self.assertEqual(analisis_count, 2)

        data_por_nivel = {n["nivel"]["id"]: n for n in niveles_data}
        for nivel_id in ["prescolar", "primaria"]:
            self.assertIn(nivel_id, data_por_nivel)
            self.assertIn("filas", data_por_nivel[nivel_id])
            self.assertIn("totales", data_por_nivel[nivel_id])
            self.assertIn("requerimientos", data_por_nivel[nivel_id])
            self.assertIn("porcentajes", data_por_nivel[nivel_id])
            self.assertIn("estados", data_por_nivel[nivel_id])
            self.assertIn("id_analisis", data_por_nivel[nivel_id])

        fila_prescolar = data_por_nivel["prescolar"]["filas"][0]
        fila_primaria = data_por_nivel["primaria"]["filas"][0]
        self.assertEqual(fila_prescolar["minimo"], 120.0)
        self.assertEqual(fila_prescolar["maximo"], 160.0)
        self.assertEqual(fila_primaria["minimo"], 180.0)
        self.assertEqual(fila_primaria["maximo"], 220.0)

    def test_api_guardar_ingredientes_por_nivel_guarda_registros_y_recalcula_totales(self):
        analisis = TablaAnalisisNutricionalMenu.objects.create(
            id_menu=self.menu,
            id_nivel_escolar_uapa=self.nivel_prescolar,
        )

        url = reverse("nutricion:api_guardar_ingredientes_por_nivel", args=[self.menu.id_menu])
        payload = {
            "niveles": [
                {
                    "id_nivel_escolar": self.nivel_prescolar.id_grado_escolar_uapa,
                    "id_analisis": analisis.id_analisis,
                    "ingredientes": [
                        {
                            "id_preparacion": self.preparacion.id_preparacion,
                            "id_ingrediente": self.alimento_icbf.codigo,
                            "peso_neto": 150.0,
                        }
                    ],
                }
            ]
        }
        response = self.client.post(url, data=json.dumps(payload), content_type="application/json")

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["success"])
        self.assertEqual(body["registros_actualizados"], 1)
        self.assertEqual(body["errores"], [])

        registro = TablaIngredientesPorNivel.objects.get(
            id_analisis=analisis,
            id_preparacion=self.preparacion,
            id_ingrediente_siesa=self.ingrediente_siesa,
        )
        self.assertAlmostEqual(float(registro.peso_neto), 150.0, places=1)
        self.assertAlmostEqual(float(registro.calorias), 90.0, places=1)

        analisis.refresh_from_db()
        self.assertAlmostEqual(float(analisis.total_calorias), 90.0, places=1)
        self.assertGreater(float(analisis.total_proteina), 0.0)

    def test_api_guardar_ingredientes_por_nivel_rechaza_metodo_get(self):
        url = reverse("nutricion:api_guardar_ingredientes_por_nivel", args=[self.menu.id_menu])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 405)

    def test_api_guardar_ingredientes_por_nivel_valida_payload_vacio(self):
        url = reverse("nutricion:api_guardar_ingredientes_por_nivel", args=[self.menu.id_menu])
        response = self.client.post(url, data=json.dumps({}), content_type="application/json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("error", response.json())

    def test_api_guardar_ingredientes_por_nivel_reporta_analisis_no_encontrado(self):
        url = reverse("nutricion:api_guardar_ingredientes_por_nivel", args=[self.menu.id_menu])
        payload = {
            "niveles": [
                {
                    "id_nivel_escolar": self.nivel_prescolar.id_grado_escolar_uapa,
                    "id_analisis": 999999,
                    "ingredientes": [
                        {
                            "id_preparacion": self.preparacion.id_preparacion,
                            "id_ingrediente": self.alimento_icbf.codigo,
                            "peso_neto": 140.0,
                        }
                    ],
                }
            ]
        }
        response = self.client.post(url, data=json.dumps(payload), content_type="application/json")

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["success"])
        self.assertEqual(body["registros_actualizados"], 0)
        self.assertTrue(any("no encontrado" in err.lower() for err in body["errores"]))

    def test_api_guardar_ingredientes_por_nivel_limita_porcentajes_para_evitar_overflow(self):
        analisis = TablaAnalisisNutricionalMenu.objects.create(
            id_menu=self.menu,
            id_nivel_escolar_uapa=self.nivel_prescolar,
        )

        url = reverse("nutricion:api_guardar_ingredientes_por_nivel", args=[self.menu.id_menu])
        payload = {
            "niveles": [
                {
                    "id_nivel_escolar": self.nivel_prescolar.id_grado_escolar_uapa,
                    "id_analisis": analisis.id_analisis,
                    "ingredientes": [
                        {
                            "id_preparacion": self.preparacion.id_preparacion,
                            "id_ingrediente": self.alimento_icbf.codigo,
                            # Fuerza un porcentaje > 1000% sin el tope.
                            "peso_neto": 15000.0,
                        }
                    ],
                }
            ]
        }
        response = self.client.post(url, data=json.dumps(payload), content_type="application/json")

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["success"])

        analisis.refresh_from_db()
        self.assertEqual(float(analisis.porcentaje_hierro), 100.0)

    def test_api_guardar_preparaciones_editor_crea_nueva_con_componente(self):
        url = reverse("nutricion:api_guardar_preparaciones_editor", args=[self.menu.id_menu])
        payload = {
            "filas": [
                {
                    "id_preparacion": None,
                    "preparacion_nombre": "Nueva preparación QA",
                    "id_componente": self.componente.id_componente,
                    "id_ingrediente": self.alimento_icbf.codigo,
                    "gramaje": None,
                }
            ]
        }
        response = self.client.post(url, data=json.dumps(payload), content_type="application/json")

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["success"])
        self.assertEqual(body["guardadas"], 1)

        prep = TablaPreparaciones.objects.get(
            id_menu=self.menu,
            preparacion="Nueva preparación QA",
        )
        self.assertEqual(prep.id_componente, self.componente)
        rel = TablaPreparacionIngredientes.objects.get(
            id_preparacion=prep,
            id_ingrediente_siesa=self.alimento_icbf,
        )
        self.assertEqual(rel.id_componente, self.componente)

    def test_api_guardar_preparaciones_editor_existente_actualiza_componente_en_preparacion_e_ingrediente(self):
        componente_alt = ComponentesAlimentos.objects.create(
            id_componente="comp_alt",
            componente="Componente Alterno",
            id_grupo_alimentos=self.grupo,
        )
        url = reverse("nutricion:api_guardar_preparaciones_editor", args=[self.menu.id_menu])
        payload = {
            "filas": [
                {
                    "id_preparacion": self.preparacion.id_preparacion,
                    "preparacion_nombre": "",
                    "id_componente": componente_alt.id_componente,
                    "id_ingrediente": self.alimento_icbf.codigo,
                    "gramaje": None,
                }
            ]
        }
        response = self.client.post(url, data=json.dumps(payload), content_type="application/json")

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["success"])

        self.preparacion.refresh_from_db()
        self.assertEqual(self.preparacion.id_componente, componente_alt)
        rel = TablaPreparacionIngredientes.objects.get(
            id_preparacion=self.preparacion,
            id_ingrediente_siesa=self.alimento_icbf,
        )
        self.assertEqual(rel.id_componente, componente_alt)

    def test_paso4_template_incluye_slider_y_labels_de_rango(self):
        url = reverse("nutricion:preparaciones_editor", args=[self.menu.id_menu])
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="slider-peso"', html=False)
        self.assertContains(response, 'class="slider-label-min"', html=False)
        self.assertContains(response, 'class="slider-label-max"', html=False)
        self.assertContains(response, 'data-minimo="', html=False)
        self.assertContains(response, 'data-maximo="', html=False)

        niveles_data = response.context["niveles_data"]
        fila_prescolar = next(n for n in niveles_data if n["nivel"]["id"] == "prescolar")["filas"][0]
        self.assertEqual(fila_prescolar["minimo"], 120.0)
        self.assertEqual(fila_prescolar["maximo"], 160.0)

    def test_paso4_template_incluye_tooltips_y_data_nutricional(self):
        url = reverse("nutricion:preparaciones_editor", args=[self.menu.id_menu])
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-bs-toggle="tooltip"', html=False)
        self.assertContains(response, 'class="input-peso"', html=False)
        self.assertContains(response, 'data-calorias="', html=False)
        self.assertContains(response, 'data-proteina="', html=False)
        self.assertContains(response, 'title="Calor', html=False)

    def test_excel_export_pobla_nutrientes_de_ingrediente_con_fallback(self):
        excel_stream = generate_menu_excel(self.menu.id_menu)
        wb = load_workbook(filename=BytesIO(excel_stream.getvalue()))
        ws = wb[wb.sheetnames[0]]

        valor_kcal = ws.cell(row=11, column=8).value
        self.assertIsNotNone(valor_kcal)
        self.assertGreater(float(valor_kcal), 0.0)

    def test_excel_export_mapea_totales_desde_claves_del_servicio(self):
        excel_stream = generate_menu_excel(self.menu.id_menu)
        wb = load_workbook(filename=BytesIO(excel_stream.getvalue()))
        ws = wb[wb.sheetnames[0]]

        total_row = None
        for row in range(11, 60):
            if ws.cell(row=row, column=1).value == "TOTAL MENÚ":
                total_row = row
                break
        self.assertIsNotNone(total_row)
        self.assertAlmostEqual(float(ws.cell(row=total_row, column=8).value), 60.0, places=1)
        self.assertAlmostEqual(float(ws.cell(row=total_row, column=9).value), 3.2, places=1)

    def test_excel_export_usa_requerimientos_y_nivel_correcto_por_hoja(self):
        TablaRequerimientosNutricionales.objects.filter(
            id_nivel_escolar_uapa=self.nivel_prescolar,
            id_modalidad=self.modalidad,
        ).update(calorias_kcal=Decimal("276.0"))
        TablaRequerimientosNutricionales.objects.filter(
            id_nivel_escolar_uapa=self.nivel_primaria,
            id_modalidad=self.modalidad,
        ).update(calorias_kcal=Decimal("417.0"))

        excel_stream = generate_menu_excel(self.menu.id_menu)
        wb = load_workbook(filename=BytesIO(excel_stream.getvalue()))

        req_por_hoja = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self.assertEqual(ws["D7"].value, sheet_name.upper())

            total_row = None
            for row in range(11, 60):
                if ws.cell(row=row, column=1).value == "TOTAL MENÚ":
                    total_row = row
                    break
            self.assertIsNotNone(total_row)
            req_row = total_row + 1
            req_por_hoja[sheet_name] = float(ws.cell(row=req_row, column=8).value)

        self.assertIn("Preescolar", req_por_hoja)
        self.assertIn("Primaria 1-3", req_por_hoja)
        self.assertEqual(req_por_hoja["Preescolar"], 276.0)
        self.assertEqual(req_por_hoja["Primaria 1-3"], 417.0)

    def test_excel_export_campos_administrativos_modalidad_y_tipo_complemento(self):
        excel_stream = generate_menu_excel(self.menu.id_menu)
        wb = load_workbook(filename=BytesIO(excel_stream.getvalue()))
        ws = wb[wb.sheetnames[0]]

        self.assertEqual(ws["D5"].value, "RACIÓN PARA PREPARAR EN SITIO")
        self.assertEqual(ws["D6"].value, self.modalidad.modalidad.upper())

    def test_excel_export_muestra_componente_combinado_por_preparacion(self):
        grupo_verduras = GruposAlimentos.objects.create(
            id_grupo_alimentos="grp_ver",
            grupo_alimentos="Verduras",
        )
        componente_verduras = ComponentesAlimentos.objects.create(
            id_componente="comp_ver",
            componente="Verduras",
            id_grupo_alimentos=grupo_verduras,
        )
        tomate = TablaAlimentos2018Icbf.objects.create(
            codigo="A002",
            nombre_del_alimento="Tomate",
            humedad_g=Decimal("94.50"),
            energia_kcal=Decimal("18.00"),
            energia_kj=Decimal("75.00"),
            proteina_g=Decimal("0.90"),
            lipidos_g=Decimal("0.20"),
            carbohidratos_totales_g=Decimal("3.90"),
            calcio_mg=Decimal("10.00"),
            hierro_mg=Decimal("0.30"),
            sodio_mg=Decimal("5.00"),
            parte_comestible_field=Decimal("95.00"),
            id_componente=componente_verduras,
        )
        TablaIngredientesSiesa.objects.create(
            id_ingrediente_siesa="A002",
            nombre_ingrediente="Tomate",
        )
        TablaPreparacionIngredientes.objects.create(
            id_preparacion=self.preparacion,
            id_ingrediente_siesa=tomate,
            id_componente=componente_verduras,
            gramaje=Decimal("20.00"),
        )

        excel_stream = generate_menu_excel(self.menu.id_menu)
        wb = load_workbook(filename=BytesIO(excel_stream.getvalue()))
        ws = wb[wb.sheetnames[0]]

        self.assertEqual(str(ws.cell(row=11, column=1).value or ""), "Bebida con leche")
        self.assertIsNone(ws.cell(row=12, column=1).value)
        self.assertIn("A11:A12", {str(rng) for rng in ws.merged_cells.ranges})
        self.assertEqual(str(ws.cell(row=11, column=2).value or ""), "Lacteos")
        self.assertEqual(str(ws.cell(row=12, column=2).value or ""), "Verduras")

    def test_servicio_aplica_ingrediente_guardado_por_codigo_icbf_sin_siesa(self):
        analisis = TablaAnalisisNutricionalMenu.objects.create(
            id_menu=self.menu,
            id_nivel_escolar_uapa=self.nivel_prescolar,
        )
        TablaIngredientesPorNivel.objects.create(
            id_analisis=analisis,
            id_preparacion=self.preparacion,
            id_preparacion_ingrediente=TablaPreparacionIngredientes.objects.get(
                id_preparacion=self.preparacion,
                id_ingrediente_siesa=self.alimento_icbf,
            ),
            id_ingrediente_siesa=None,
            codigo_icbf=self.alimento_icbf.codigo,
            peso_neto=Decimal("222.00"),
            peso_bruto=Decimal("222.00"),
            calorias=Decimal("123.00"),
            proteina=Decimal("11.00"),
            grasa=Decimal("10.00"),
            cho=Decimal("9.00"),
            calcio=Decimal("8.00"),
            hierro=Decimal("7.00"),
            sodio=Decimal("6.00"),
        )

        resultado = AnalisisNutricionalService.obtener_analisis_completo(self.menu.id_menu)
        self.assertTrue(resultado["success"])
        nivel_prescolar = next(
            n for n in resultado["analisis_por_nivel"]
            if n["nivel_escolar"]["id"] == self.nivel_prescolar.id_grado_escolar_uapa
        )
        ingrediente = nivel_prescolar["preparaciones"][0]["ingredientes"][0]

        self.assertEqual(float(ingrediente["peso_neto_base"]), 222.0)
        self.assertIn("valores_finales_guardados", ingrediente)
        self.assertEqual(float(ingrediente["valores_finales_guardados"]["calorias"]), 123.0)

    def test_excel_export_usa_firmas_configuradas_por_contrato(self):
        FirmaNutricionalContrato.objects.create(
            programa=self.programa,
            elabora_nombre="Nutri Elabora QA",
            elabora_matricula="MAT-111",
            elabora_firma_texto="FIRMA ELABORA QA",
            aprueba_nombre="Nutri Aprueba QA",
            aprueba_matricula="MAT-222",
            aprueba_firma_texto="FIRMA APRUEBA QA",
        )

        excel_stream = generate_menu_excel(self.menu.id_menu)
        wb = load_workbook(filename=BytesIO(excel_stream.getvalue()))
        ws = wb[wb.sheetnames[0]]

        row_elabora = None
        row_aprueba = None
        row_matriculas = []
        for row in range(10, 80):
            col1 = ws.cell(row=row, column=1).value
            col8 = ws.cell(row=row, column=8).value
            if col1 == "NOMBRE NUTRICIONISTA - DIETISTA QUE ELABORA EL ANÁLISIS":
                row_elabora = row
            if col1 == "NOMBRE NUTRICIONISTA - DIETISTA QUE REVISA Y APRUEBA EL ANÁLISIS POR PARTE DE LA SEM":
                row_aprueba = row
            if col8 == "MATRÍCULA PROFESIONAL":
                row_matriculas.append(row)

        self.assertIsNotNone(row_elabora)
        self.assertIsNotNone(row_aprueba)
        self.assertEqual(ws.cell(row=row_elabora, column=5).value, "NUTRI ELABORA QA")
        self.assertEqual(ws.cell(row=row_aprueba, column=5).value, "NUTRI APRUEBA QA")
        self.assertGreaterEqual(len(row_matriculas), 2)
        self.assertEqual(ws.cell(row=row_matriculas[0], column=11).value, "MAT-111")
        self.assertEqual(ws.cell(row=row_matriculas[1], column=11).value, "MAT-222")

    def test_vista_firmas_contrato_carga(self):
        url = reverse("nutricion:firmas_contrato")
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Firmas Nutricionales por Contrato", html=False)


class Paso4FrontendContractsTests(TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        base_dir = Path(__file__).resolve().parents[1]
        js_path = base_dir / "static" / "js" / "nutricion" / "preparaciones_editor.js"
        cls.js_source = js_path.read_text(encoding="utf-8")

    def test_paso4_js_contiene_funciones_clave_de_sync_y_validacion(self):
        js = self.js_source
        self.assertIn("function sincronizarSliderConInput(row)", js)
        self.assertIn("function sincronizarInputConSlider(row)", js)
        self.assertIn("function actualizarEstadoFila(row)", js)
        self.assertIn("if (e.target.classList.contains('input-peso'))", js)
        self.assertIn("if (e.target.classList.contains('slider-peso'))", js)
        self.assertIn("badge.textContent = resultado.valido ? 'OK' : 'FUERA';", js)

    def test_paso4_js_contiene_overlay_y_uso_en_guardar(self):
        js = self.js_source
        self.assertIn("function mostrarOverlayGuardando(mensaje = 'Guardando cambios...')", js)
        self.assertIn("function ocultarOverlayGuardando()", js)
        self.assertIn("const overlay = mostrarOverlayGuardando('Guardando cambios...');", js)
        self.assertIn("ocultarOverlayGuardando();", js)

    def test_paso4_js_modal_componente_visible_y_requerido(self):
        js = self.js_source
        self.assertIn("bComponente.style.display = 'block';", js)
        self.assertIn("selectComponente.disabled = false;", js)
        self.assertIn("if (!idComp)", js)
        self.assertIn("id_componente: idComp || null", js)


class ExcelRulesTests(TestCase):
    def test_resolver_modalidad_atencion(self):
        self.assertEqual(
            ExcelReportDrawer._resolver_modalidad_atencion("020511"),
            "Ración Industrializada"
        )
        self.assertEqual(
            ExcelReportDrawer._resolver_modalidad_atencion("20502"),
            "Ración Industrializada"
        )
        self.assertEqual(
            ExcelReportDrawer._resolver_modalidad_atencion("modpaso2"),
            "Ración para Preparar en Sitio"
        )


class CopiaPreparacionApiTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_user(username="copy_prep_user", password="testpass123")

        cls.modalidad_a = ModalidadesDeConsumo.objects.create(
            id_modalidades="mod_copy_a",
            modalidad="CAJM AM Copy A",
            cod_modalidad="CAJM_A",
        )
        cls.modalidad_b = ModalidadesDeConsumo.objects.create(
            id_modalidades="mod_copy_b",
            modalidad="CAJM AM Copy B",
            cod_modalidad="CAJM_B",
        )
        cls.municipio = PrincipalMunicipio.objects.create(
            codigo_municipio=22222,
            nombre_municipio="Municipio Copy",
            codigo_departamento="22",
        )
        cls.programa_a = Programa.objects.create(
            programa="Programa Copy A",
            contrato="CT-COPY-A",
            municipio=cls.municipio,
            fecha_inicial=date(2025, 1, 1),
            fecha_final=date(2025, 12, 31),
            estado="activo",
        )
        cls.programa_b = Programa.objects.create(
            programa="Programa Copy B",
            contrato="CT-COPY-B",
            municipio=cls.municipio,
            fecha_inicial=date(2025, 1, 1),
            fecha_final=date(2025, 12, 31),
            estado="activo",
        )

        cls.menu_origen = TablaMenus.objects.create(
            menu="Menu Origen Copy",
            id_modalidad=cls.modalidad_a,
            id_contrato=cls.programa_a,
        )
        cls.menu_destino_misma_modalidad = TablaMenus.objects.create(
            menu="Menu Destino Copy A",
            id_modalidad=cls.modalidad_a,
            id_contrato=cls.programa_a,
        )
        cls.menu_destino_otra_modalidad = TablaMenus.objects.create(
            menu="Menu Destino Copy B",
            id_modalidad=cls.modalidad_b,
            id_contrato=cls.programa_b,
        )

        cls.grupo = GruposAlimentos.objects.create(
            id_grupo_alimentos="grp_copy",
            grupo_alimentos="Frutas",
        )
        cls.componente = ComponentesAlimentos.objects.create(
            id_componente="comp_copy",
            componente="Fruta fresca",
            id_grupo_alimentos=cls.grupo,
        )

        cls.alimento_a = TablaAlimentos2018Icbf.objects.create(
            codigo="CP01",
            nombre_del_alimento="Manzana",
            humedad_g=Decimal("85.00"),
            energia_kcal=52,
            energia_kj=218,
            proteina_g=Decimal("0.30"),
            lipidos_g=Decimal("0.20"),
            carbohidratos_totales_g=Decimal("14.00"),
            calcio_mg=6,
            hierro_mg=Decimal("0.10"),
            sodio_mg=1,
            parte_comestible_field=95,
            id_componente=cls.componente,
        )
        cls.alimento_b = TablaAlimentos2018Icbf.objects.create(
            codigo="CP02",
            nombre_del_alimento="Banano",
            humedad_g=Decimal("74.00"),
            energia_kcal=89,
            energia_kj=372,
            proteina_g=Decimal("1.10"),
            lipidos_g=Decimal("0.30"),
            carbohidratos_totales_g=Decimal("23.00"),
            calcio_mg=5,
            hierro_mg=Decimal("0.30"),
            sodio_mg=1,
            parte_comestible_field=70,
            id_componente=cls.componente,
        )

        cls.preparacion_origen = TablaPreparaciones.objects.create(
            preparacion="Fruta mixta copy",
            id_menu=cls.menu_origen,
            id_componente=cls.componente,
        )
        TablaPreparacionIngredientes.objects.create(
            id_preparacion=cls.preparacion_origen,
            id_ingrediente_siesa=cls.alimento_a,
            id_componente=cls.componente,
            id_grupo_alimentos=cls.grupo,
            gramaje=Decimal("80.00"),
        )
        TablaPreparacionIngredientes.objects.create(
            id_preparacion=cls.preparacion_origen,
            id_ingrediente_siesa=cls.alimento_b,
            id_componente=cls.componente,
            id_grupo_alimentos=cls.grupo,
            gramaje=Decimal("60.00"),
        )

    def setUp(self):
        self.client = Client()
        self.client.force_login(self.user)

    def test_api_copiar_preparacion_permite_copia_parcial_por_ingredientes(self):
        url = reverse("nutricion:api_copiar_preparacion")
        payload = {
            "source_preparacion_id": self.preparacion_origen.id_preparacion,
            "target_menu_id": self.menu_destino_misma_modalidad.id_menu,
            "ingredient_ids": ["CP01"],
        }

        response = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["success"])

        nueva_id = body["nueva_preparacion"]["id_preparacion"]
        nueva_prep = TablaPreparaciones.objects.get(id_preparacion=nueva_id)
        ingredientes = list(
            TablaPreparacionIngredientes.objects.filter(id_preparacion=nueva_prep)
            .values_list("id_ingrediente_siesa_id", flat=True)
        )
        self.assertEqual(ingredientes, ["CP01"])

    def test_api_copiar_preparacion_rechaza_modalidad_distinta(self):
        url = reverse("nutricion:api_copiar_preparacion")
        payload = {
            "source_preparacion_id": self.preparacion_origen.id_preparacion,
            "target_menu_id": self.menu_destino_otra_modalidad.id_menu,
        }

        response = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(response.status_code, 400)
        body = response.json()
        self.assertFalse(body["success"])
        self.assertIn("misma modalidad", body["error"])
