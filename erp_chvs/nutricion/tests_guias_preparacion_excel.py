from datetime import date
from decimal import Decimal
from io import BytesIO

from django.test import TestCase
from openpyxl import load_workbook

from planeacion.models import Programa
from principal.models import ModalidadesDeConsumo, PrincipalMunicipio, TablaGradosEscolaresUapa

from .guia_preparacion_excel_generator import GuiaPreparacionExcelGenerator
from .models import (
    ComponentesAlimentos,
    FirmaNutricionalContrato,
    GruposAlimentos,
    TablaAlimentos2018Icbf,
    TablaAnalisisNutricionalMenu,
    TablaIngredientesPorNivel,
    TablaIngredientesSiesa,
    TablaMenus,
    TablaPreparacionIngredientes,
    TablaPreparaciones,
)


class GuiasPreparacionExcelTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.nivel_prescolar = TablaGradosEscolaresUapa.objects.create(
            id_grado_escolar_uapa="100",
            nivel_escolar_uapa="prescolar",
        )

        cls.modalidad = ModalidadesDeConsumo.objects.create(
            id_modalidades="mod_guia",
            modalidad="COMPLEMENTO AM/PM PREPARADO",
            cod_modalidad="CAJM",
        )
        cls.municipio = PrincipalMunicipio.objects.create(
            codigo_municipio=22222,
            nombre_municipio="Municipio Guia",
            codigo_departamento="22",
        )
        cls.programa = Programa.objects.create(
            programa="Programa Guia",
            contrato="CT-GUIA-001",
            municipio=cls.municipio,
            fecha_inicial=date(2026, 1, 1),
            fecha_final=date(2026, 12, 31),
            estado="activo",
        )

        cls.grupo = GruposAlimentos.objects.create(
            id_grupo_alimentos="grp_guia",
            grupo_alimentos="Cereales",
        )
        cls.componente = ComponentesAlimentos.objects.create(
            id_componente="comp_guia",
            componente="Bebida",
            id_grupo_alimentos=cls.grupo,
        )

        cls.alimento_1 = TablaAlimentos2018Icbf.objects.create(
            codigo="G001",
            nombre_del_alimento="Ingrediente Guia 1",
            humedad_g=Decimal("1.00"),
            energia_kcal=Decimal("100.00"),
            energia_kj=Decimal("100.00"),
            proteina_g=Decimal("1.00"),
            lipidos_g=Decimal("1.00"),
            carbohidratos_totales_g=Decimal("1.00"),
            calcio_mg=Decimal("1.00"),
            hierro_mg=Decimal("1.00"),
            sodio_mg=Decimal("1.00"),
            parte_comestible_field=Decimal("100.00"),
            id_componente=cls.componente,
        )
        cls.alimento_2 = TablaAlimentos2018Icbf.objects.create(
            codigo="G002",
            nombre_del_alimento="Ingrediente Guia 2",
            humedad_g=Decimal("1.00"),
            energia_kcal=Decimal("100.00"),
            energia_kj=Decimal("100.00"),
            proteina_g=Decimal("1.00"),
            lipidos_g=Decimal("1.00"),
            carbohidratos_totales_g=Decimal("1.00"),
            calcio_mg=Decimal("1.00"),
            hierro_mg=Decimal("1.00"),
            sodio_mg=Decimal("1.00"),
            parte_comestible_field=Decimal("100.00"),
            id_componente=cls.componente,
        )
        TablaIngredientesSiesa.objects.create(id_ingrediente_siesa="G001", nombre_ingrediente="I1")
        TablaIngredientesSiesa.objects.create(id_ingrediente_siesa="G002", nombre_ingrediente="I2")

        cls.menu_1 = TablaMenus.objects.create(
            menu="1",
            id_modalidad=cls.modalidad,
            id_contrato=cls.programa,
        )
        cls.menu_2 = TablaMenus.objects.create(
            menu="2",
            id_modalidad=cls.modalidad,
            id_contrato=cls.programa,
        )

        cls.prep_1 = TablaPreparaciones.objects.create(
            preparacion="Preparacion Guia",
            id_menu=cls.menu_1,
            id_componente=cls.componente,
        )
        cls.prep_1_ing_1 = TablaPreparacionIngredientes.objects.create(
            id_preparacion=cls.prep_1,
            id_ingrediente_siesa=cls.alimento_1,
            gramaje=Decimal("10.00"),
        )
        cls.prep_1_ing_2 = TablaPreparacionIngredientes.objects.create(
            id_preparacion=cls.prep_1,
            id_ingrediente_siesa=cls.alimento_2,
            gramaje=Decimal("20.00"),
        )

        cls.prep_2 = TablaPreparaciones.objects.create(
            preparacion="Preparacion Guia 2",
            id_menu=cls.menu_2,
            id_componente=cls.componente,
        )
        cls.prep_2_ing_1 = TablaPreparacionIngredientes.objects.create(
            id_preparacion=cls.prep_2,
            id_ingrediente_siesa=cls.alimento_1,
            gramaje=Decimal("15.00"),
        )

        analisis_1 = TablaAnalisisNutricionalMenu.objects.create(
            id_menu=cls.menu_1,
            id_nivel_escolar_uapa=cls.nivel_prescolar,
        )
        TablaIngredientesPorNivel.objects.create(
            id_analisis=analisis_1,
            id_preparacion=cls.prep_1,
            id_preparacion_ingrediente=cls.prep_1_ing_1,
            codigo_icbf="G001",
            peso_neto=Decimal("10.00"),
            peso_bruto=Decimal("10.00"),
        )
        TablaIngredientesPorNivel.objects.create(
            id_analisis=analisis_1,
            id_preparacion=cls.prep_1,
            id_preparacion_ingrediente=cls.prep_1_ing_2,
            codigo_icbf="G002",
            peso_neto=Decimal("20.00"),
            peso_bruto=Decimal("20.00"),
        )

    def test_genera_una_hoja_por_menu(self):
        generator = GuiaPreparacionExcelGenerator()
        stream = generator.generate(self.programa.id, self.modalidad.id_modalidades)
        wb = load_workbook(filename=BytesIO(stream.getvalue()))
        self.assertIn("Menu 1", wb.sheetnames)
        self.assertIn("Menu 2", wb.sheetnames)
        self.assertEqual(len(wb.sheetnames), 2)

    def test_peso_servido_es_suma_de_netos_por_preparacion(self):
        generator = GuiaPreparacionExcelGenerator()
        stream = generator.generate(self.programa.id, self.modalidad.id_modalidades)
        wb = load_workbook(filename=BytesIO(stream.getvalue()))
        ws = wb["Menu 1"]

        # Primera fila de datos: prescolar -> C bruto, D neto, E servido
        servido_fila_1 = ws.cell(row=11, column=5).value
        servido_fila_2 = ws.cell(row=12, column=5).value
        self.assertEqual(float(servido_fila_1), 30.0)
        self.assertIsNone(servido_fila_2)

    def test_toma_pesos_guardados_con_ids_nivel_numericos_y_gramaje_nulo(self):
        TablaPreparacionIngredientes.objects.filter(id_preparacion=self.prep_1).update(gramaje=None)

        generator = GuiaPreparacionExcelGenerator()
        stream = generator.generate(self.programa.id, self.modalidad.id_modalidades)
        wb = load_workbook(filename=BytesIO(stream.getvalue()))
        ws = wb["Menu 1"]

        # Fila 11 = primer ingrediente. Prescolar ocupa C:D:E = bruto, neto, servido.
        self.assertEqual(float(ws.cell(row=11, column=3).value), 10.0)
        self.assertEqual(float(ws.cell(row=11, column=4).value), 10.0)
        self.assertEqual(float(ws.cell(row=11, column=5).value), 30.0)

    def test_bloque_firmas_usa_tabla_firma_nutricional_contrato(self):
        FirmaNutricionalContrato.objects.create(
            programa=self.programa,
            elabora_nombre="Dietista Elabora",
            elabora_matricula="MAT-ELA-1",
            elabora_firma_texto="FIRMA ELA",
            aprueba_nombre="Dietista Aprueba",
            aprueba_matricula="MAT-APR-2",
            aprueba_firma_texto="FIRMA APR",
        )

        generator = GuiaPreparacionExcelGenerator()
        stream = generator.generate(self.programa.id, self.modalidad.id_modalidades)
        wb = load_workbook(filename=BytesIO(stream.getvalue()))
        ws = wb["Menu 1"]

        found_elabora = False
        found_aprueba = False
        for row in range(1, 200):
            text = ws.cell(row=row, column=1).value
            if text == "NOMBRE NUTRICIONISTA - DIETISTA POR PARTE DEL OPERADOR":
                self.assertEqual(ws.cell(row=row, column=5).value, "DIETISTA ELABORA")
                self.assertEqual(ws.cell(row=row, column=12).value, "MAT-ELA-1")
                found_elabora = True
            if text == "NOMBRE NUTRICIONISTA - DIETISTA QUE APRUEBA LA GUIA POR PARTE DEL PAE SEM":
                self.assertEqual(ws.cell(row=row, column=5).value, "DIETISTA APRUEBA")
                self.assertEqual(ws.cell(row=row, column=12).value, "MAT-APR-2")
                found_aprueba = True

        self.assertTrue(found_elabora)
        self.assertTrue(found_aprueba)

    def test_pestanas_se_ordenan_numericamente(self):
        TablaMenus.objects.create(
            menu="10",
            id_modalidad=self.modalidad,
            id_contrato=self.programa,
        )

        generator = GuiaPreparacionExcelGenerator()
        stream = generator.generate(self.programa.id, self.modalidad.id_modalidades)
        wb = load_workbook(filename=BytesIO(stream.getvalue()))

        self.assertEqual(wb.sheetnames[:3], ["Menu 1", "Menu 2", "Menu 10"])

    def test_no_hay_filas_vacias_entre_tabla_y_bloque_firmas(self):
        FirmaNutricionalContrato.objects.create(
            programa=self.programa,
            elabora_nombre="Dietista Elabora",
            elabora_matricula="MAT-ELA-1",
            aprueba_nombre="Dietista Aprueba",
            aprueba_matricula="MAT-APR-2",
        )

        generator = GuiaPreparacionExcelGenerator()
        stream = generator.generate(self.programa.id, self.modalidad.id_modalidades)
        wb = load_workbook(filename=BytesIO(stream.getvalue()))
        ws = wb["Menu 1"]

        # Ultima fila de ingredientes para menu 1 en setup: fila 12.
        # Bloque firmas debe iniciar inmediatamente en la fila 13.
        self.assertEqual(
            ws.cell(row=13, column=1).value,
            "NOMBRE NUTRICIONISTA - DIETISTA POR PARTE DEL OPERADOR",
        )
