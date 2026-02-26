from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from nutricion.models import (
    TablaMenus, TablaPreparaciones, TablaPreparacionIngredientes, 
    TablaIngredientesPorNivel, TablaAlimentos2018Icbf
)
from principal.models import PrincipalMunicipio, TablaGradosEscolaresUapa, ModalidadesDeConsumo, RegistroActividad
from planeacion.models import Programa
from django.db.models import Q
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

@login_required
def costos_index(request):
    """
    Vista principal del módulo de costos.
    """
    return render(request, 'costos/index.html')

def get_filtered_matriz_queryset(request):
    """
    Función de utilidad para obtener el queryset filtrado compartido por la vista y el exportador.
    """
    municipio_id = request.GET.get('municipio')
    programa_id = request.GET.get('programa')
    modalidad_id = request.GET.get('modalidad')
    semana = request.GET.get('semana')
    menu_nombre = request.GET.get('menu')
    preparacion_nombre = request.GET.get('preparacion')

    queryset = TablaIngredientesPorNivel.objects.select_related(
        'id_analisis__id_menu__id_modalidad',
        'id_analisis__id_menu__id_contrato',
        'id_analisis__id_nivel_escolar_uapa',
        'id_preparacion',
    ).all()

    if municipio_id:
        queryset = queryset.filter(id_analisis__id_menu__id_contrato__municipio_id=municipio_id)
    
    if programa_id:
        queryset = queryset.filter(id_analisis__id_menu__id_contrato_id=programa_id)
        
    if modalidad_id:
        queryset = queryset.filter(id_analisis__id_menu__id_modalidad_id=modalidad_id)
        
    if semana:
        queryset = queryset.filter(id_analisis__id_menu__semana=semana)
        
    if menu_nombre:
        queryset = queryset.filter(id_analisis__id_menu__menu__icontains=menu_nombre)
        
    if preparacion_nombre:
        queryset = queryset.filter(id_preparacion__preparacion__icontains=preparacion_nombre)

    return queryset.order_by(
        'id_analisis__id_menu__id_modalidad__modalidad',
        'id_analisis__id_nivel_escolar_uapa__nivel_escolar_uapa',
        'id_analisis__id_menu__semana',
        'id_analisis__id_menu__menu',
        'id_preparacion__preparacion'
    )

@login_required
def matriz_nutricional(request):
    """
    Vista de la Matriz Nutricional con información de ingredientes, pesos y niveles.
    """
    # Parámetros de filtro para el contexto de la plantilla
    municipio_id = request.GET.get('municipio')
    programa_id = request.GET.get('programa')
    modalidad_id = request.GET.get('modalidad')
    semana = request.GET.get('semana')
    menu_nombre = request.GET.get('menu')
    preparacion_nombre = request.GET.get('preparacion')
    
    # 1. Municipios que tienen al menos un contrato/programa
    municipios_con_contrato_ids = Programa.objects.values_list('municipio_id', flat=True).distinct()
    municipios = PrincipalMunicipio.objects.filter(id__in=municipios_con_contrato_ids).order_by('nombre_municipio')
    
    # 2. Programas filtrados por municipio si aplica
    programas = Programa.objects.all().order_by('programa')
    if municipio_id:
        programas = programas.filter(municipio_id=municipio_id)

    # 3. Modalidades
    modalidades = ModalidadesDeConsumo.objects.all().order_by('modalidad')

    # Obtener datos filtrados
    queryset = get_filtered_matriz_queryset(request)

    # Optimización de carga de nombres ICBF
    codigos_icbf = queryset.values_list('codigo_icbf', flat=True).distinct()
    alimentos_dict = {a.codigo: a.nombre_del_alimento for a in TablaAlimentos2018Icbf.objects.filter(codigo__in=codigos_icbf)}

    matriz_data = []
    for item in queryset:
        matriz_data.append({
            'modalidad': item.id_analisis.id_menu.id_modalidad.modalidad,
            'nivel_escolar': item.id_analisis.id_nivel_escolar_uapa.nivel_escolar_uapa,
            'semana': item.id_analisis.id_menu.semana or 'N/A',
            'menu': item.id_analisis.id_menu.menu,
            'preparacion': item.id_preparacion.preparacion,
            'ingrediente': alimentos_dict.get(item.codigo_icbf, item.codigo_icbf),
            'peso_bruto': item.peso_bruto,
            'peso_neto': item.peso_neto,
        })

    context = {
        'municipios': municipios,
        'programas': programas,
        'modalidades': modalidades,
        'semanas': [1, 2, 3, 4],
        'matriz_data': matriz_data,
        'selected_municipio': municipio_id,
        'selected_programa': programa_id,
        'selected_modalidad': modalidad_id,
        'selected_semana': semana,
        'selected_menu': menu_nombre,
        'selected_preparacion': preparacion_nombre,
    }
    
    return render(request, 'costos/matriz_nutricional.html', context)

@login_required
def exportar_matriz_excel(request):
    """
    Genera un archivo Excel con la información de la matriz nutricional filtrada.
    """
    queryset = get_filtered_matriz_queryset(request)
    programa_id = request.GET.get('programa')
    
    # Obtener el nombre del programa para el encabezado
    nombre_programa = "CONSOLIDADO DE PROGRAMAS"
    if programa_id:
        try:
            p = Programa.objects.get(id=programa_id)
            nombre_programa = f"PROGRAMA: {p.programa}"
        except Programa.DoesNotExist:
            pass
    elif queryset.exists():
        # Si no hay filtro pero hay datos, podríamos intentar ver si todos son del mismo programa
        programas_en_datos = queryset.values_list('id_analisis__id_menu__id_contrato__programa', flat=True).distinct()
        if programas_en_datos.count() == 1:
            nombre_programa = f"PROGRAMA: {programas_en_datos[0]}"

    # Optimización de carga de nombres ICBF
    codigos_icbf = queryset.values_list('codigo_icbf', flat=True).distinct()
    alimentos_dict = {a.codigo: a.nombre_del_alimento for a in TablaAlimentos2018Icbf.objects.filter(codigo__in=codigos_icbf)}

    # Crear el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Matriz Nutricional"

    # Estilos
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # Fila 1: Nombre del Programa
    ws.merge_cells('A1:H1')
    cell_title = ws['A1']
    cell_title.value = nombre_programa
    cell_title.font = title_font
    cell_title.alignment = center_align
    ws.row_dimensions[1].height = 25

    # Cabeceras (Fila 2)
    headers = [
        "MODALIDAD", "NIVEL ESCOLAR", "SEMANA", "MENU", 
        "PREPARACIÓN", "NOMBRE DEL ALIMENTO (Ingredientes)", 
        "PESO BRUTO (g)", "PESO NETO (g)"
    ]
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Datos (A partir de la Fila 3)
    for row_num, item in enumerate(queryset, 3):
        ws.cell(row=row_num, column=1, value=item.id_analisis.id_menu.id_modalidad.modalidad).border = border
        ws.cell(row=row_num, column=2, value=item.id_analisis.id_nivel_escolar_uapa.nivel_escolar_uapa).border = border
        ws.cell(row=row_num, column=3, value=item.id_analisis.id_menu.semana or 'N/A').border = border
        ws.cell(row=row_num, column=4, value=item.id_analisis.id_menu.menu).border = border
        ws.cell(row=row_num, column=5, value=item.id_preparacion.preparacion).border = border
        ws.cell(row=row_num, column=6, value=alimentos_dict.get(item.codigo_icbf, item.codigo_icbf)).border = border
        ws.cell(row=row_num, column=7, value=float(item.peso_bruto)).border = border
        ws.cell(row=row_num, column=8, value=float(item.peso_neto)).border = border

    # Ajustar ancho de columnas
    column_widths = [20, 20, 10, 10, 25, 45, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = width

    # Preparar la respuesta
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=matriz_nutricional.xlsx'
    RegistroActividad.registrar(
        request, 'costos', 'exportar_excel',
        f"Programa ID: {request.GET.get('programa', 'todos')} | Modalidad: {request.GET.get('modalidad', 'todas')}"
    )
    return response
